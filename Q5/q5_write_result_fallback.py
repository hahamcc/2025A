"""Dependency-free fallback for filling the official Question 5 workbook."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def read_rows(path: Path) -> dict[tuple[str, int], dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return {(row["uav"], int(row["bomb"])): row for row in csv.DictReader(handle)}


def main() -> None:
    parser = argparse.ArgumentParser(description="填写问题五 result3.xlsx 的备用写入器")
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--solution", type=Path, required=True)
    parser.add_argument("--contributions", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    solution = read_rows(args.solution)
    contributions = read_rows(args.contributions)
    workbook = load_workbook(args.template)
    sheet = workbook["Sheet1"]
    for row_number in range(2, 17):
        uav = str(sheet.cell(row_number, 1).value)
        bomb = int(sheet.cell(row_number, 4).value)
        key = (uav, bomb)
        row = solution.get(key)
        contribution = contributions.get(key)
        used = row is not None and contribution is not None and contribution["used"].strip().lower() == "true"
        if not used:
            for column in (*range(2, 4), *range(5, 13)):
                sheet.cell(row_number, column).value = None
            continue
        missile = contribution["primary_missile"]
        individual_key = f"{missile}_individual_s"
        values = (
            float(row["heading_deg"]),
            float(row["speed_mps"]),
            float(row["release_x_m"]),
            float(row["release_y_m"]),
            float(row["release_z_m"]),
            float(row["burst_x_m"]),
            float(row["burst_y_m"]),
            float(row["burst_z_m"]),
            float(row[individual_key]),
            missile,
        )
        for column, value in zip((*range(2, 4), *range(5, 13)), values):
            sheet.cell(row_number, column).value = value
    for row in sheet.iter_rows(min_row=2, max_row=16, min_col=2, max_col=11):
        for cell in row:
            cell.number_format = "0.000000"
    workbook.save(args.output)
    print(f"已写入：{args.output}")


if __name__ == "__main__":
    main()

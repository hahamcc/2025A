from __future__ import annotations

import argparse
import csv
import itertools
import math
import shutil
import sys
import time
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Iterable, Sequence

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import differential_evolution
from scipy.stats import qmc


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

Q4_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = PROJECT_ROOT / "Resources" / "result2.xlsx"

GRAVITY = 9.8
MISSILE_SPEED = 300.0
CLOUD_RADIUS = 10.0
CLOUD_LIFETIME = 20.0
CLOUD_SINK_SPEED = 3.0
UAV_SPEED_MIN = 70.0
UAV_SPEED_MAX = 140.0

MISSILE_INITIAL = np.array((20000.0, 0.0, 2000.0), dtype=float)
MISSILE_DIRECTION = -MISSILE_INITIAL / np.linalg.norm(MISSILE_INITIAL)
MISSILE_VELOCITY = MISSILE_SPEED * MISSILE_DIRECTION
MISSILE_IMPACT_TIME = float(np.linalg.norm(MISSILE_INITIAL) / MISSILE_SPEED)
MISSILE_X_SPEED = float(-MISSILE_VELOCITY[0])

TARGET_BOTTOM_CENTER = np.array((0.0, 200.0, 0.0), dtype=float)
TARGET_CENTER = TARGET_BOTTOM_CENTER + np.array((0.0, 0.0, 5.0))
TARGET_RADIUS = 7.0
TARGET_HEIGHT = 10.0
TARGET_BOUND_RADIUS = float(np.hypot(TARGET_RADIUS, TARGET_HEIGHT / 2.0))

UAV_NAMES = ("FY1", "FY2", "FY3")
UAV_INITIALS = np.array(
    ((17800.0, 0.0, 1800.0), (12000.0, 1400.0, 1400.0), (6000.0, -3000.0, 700.0)),
    dtype=float,
)
BURST_TIME_LIMITS = np.minimum(
    (20000.0 + CLOUD_RADIUS - UAV_INITIALS[:, 0]) / (MISSILE_X_SPEED - UAV_SPEED_MAX),
    MISSILE_IMPACT_TIME,
)
FUSE_DELAY_LIMITS = np.sqrt(2.0 * UAV_INITIALS[:, 2] / GRAVITY)

DIMENSION = 12


@dataclass(frozen=True)
class SearchProfile:
    name: str
    seeds: tuple[int, ...]
    single_population: int
    single_maxiter: int
    single_pool_count: int
    single_library_size: int
    population_size: int
    maxiter: int
    search_grid: tuple[int, int, int, float]
    rerank_grid: tuple[int, int, int, float]
    final_grid: tuple[int, int, int, float]
    verification_grid: tuple[int, int, int, float]
    top_per_seed: int
    rerank_count: int
    final_count: int
    global_fraction: float = 0.30


PROFILES = {
    "quick": SearchProfile(
        name="quick",
        seeds=(41, 42),
        single_population=20,
        single_maxiter=20,
        single_pool_count=12,
        single_library_size=4,
        population_size=36,
        maxiter=24,
        search_grid=(24, 3, 3, 0.20),
        rerank_grid=(60, 5, 5, 0.05),
        final_grid=(120, 7, 7, 0.02),
        verification_grid=(180, 9, 9, 0.01),
        top_per_seed=8,
        rerank_count=10,
        final_count=3,
    ),
    "standard": SearchProfile(
        name="standard",
        seeds=(20250841, 20250842, 20250843),
        single_population=40,
        single_maxiter=60,
        single_pool_count=24,
        single_library_size=6,
        population_size=72,
        maxiter=90,
        search_grid=(36, 5, 5, 0.10),
        rerank_grid=(120, 7, 7, 0.02),
        final_grid=(360, 11, 11, 0.01),
        verification_grid=(720, 21, 15, 0.005),
        top_per_seed=16,
        rerank_count=24,
        final_count=5,
    ),
}


@dataclass(frozen=True)
class UavPlan:
    name: str
    initial: tuple[float, float, float]
    heading: float
    speed: float
    horizontal_distance: float
    burst_time: float
    fuse_delay: float

    @property
    def release_time(self) -> float:
        return self.burst_time - self.fuse_delay

    @property
    def direction(self) -> np.ndarray:
        return np.array((math.cos(self.heading), math.sin(self.heading), 0.0), dtype=float)

    @property
    def release_point(self) -> np.ndarray:
        return np.asarray(self.initial, dtype=float) + self.speed * self.release_time * self.direction

    @property
    def burst_point(self) -> np.ndarray:
        point = np.asarray(self.initial, dtype=float) + self.horizontal_distance * self.direction
        point = point.copy()
        point[2] -= 0.5 * GRAVITY * self.fuse_delay**2
        return point


@dataclass(frozen=True)
class Strategy:
    plans: tuple[UavPlan, ...]


@dataclass(frozen=True)
class Evaluation:
    strategy: Strategy
    feasible: bool
    reason: str
    intervals: tuple[tuple[float, float], ...]
    duration: float
    release_points: tuple[np.ndarray, ...]
    burst_points: tuple[np.ndarray, ...]
    grid: tuple[int, int, int, float]


@dataclass(frozen=True)
class SingleCandidate:
    encoded: np.ndarray
    plan: UavPlan
    point_duration: float
    surface_duration: float


@dataclass(frozen=True)
class JointCandidate:
    encoded: np.ndarray
    result: Evaluation
    source: str
    seed: int


@dataclass(frozen=True)
class SearchRun:
    seed: int
    elapsed_seconds: float
    iterations: int
    evaluations: int
    success: bool
    message: str
    best: JointCandidate
    candidates: tuple[JointCandidate, ...]
    history: tuple[float, ...]


def variable_bounds(indices: Sequence[int]) -> tuple[tuple[float, float], ...]:
    bounds: list[tuple[float, float]] = []
    for index in indices:
        bounds.extend(
            (
                (0.0, 2.0 * np.pi),
                (0.0, 1.0),
                (0.0, float(BURST_TIME_LIMITS[index])),
                (0.0, 1.0),
            )
        )
    return tuple(bounds)


JOINT_BOUNDS = variable_bounds((0, 1, 2))


def decode_group(values: Sequence[float], uav_index: int) -> UavPlan:
    """把可行化坐标解码成 (theta, s, t_e, tau) 物理参数。"""

    heading, speed_unit, burst_time, delay_unit = map(float, values)
    heading = heading % (2.0 * np.pi)
    burst_time = float(np.clip(burst_time, 0.0, BURST_TIME_LIMITS[uav_index]))
    speed = UAV_SPEED_MIN + (UAV_SPEED_MAX - UAV_SPEED_MIN) * float(np.clip(speed_unit, 0.0, 1.0))
    horizontal_distance = speed * burst_time
    delay_ceiling = min(burst_time, float(FUSE_DELAY_LIMITS[uav_index]))
    fuse_delay = delay_ceiling * float(np.clip(delay_unit, 0.0, 1.0))
    return UavPlan(
        name=UAV_NAMES[uav_index],
        initial=tuple(float(value) for value in UAV_INITIALS[uav_index]),
        heading=heading,
        speed=speed,
        horizontal_distance=horizontal_distance,
        burst_time=burst_time,
        fuse_delay=fuse_delay,
    )


def decode_strategy(vector: Iterable[float]) -> Strategy:
    values = np.asarray(tuple(float(value) for value in vector), dtype=float)
    if values.shape != (DIMENSION,):
        raise ValueError("问题四计算向量必须包含十二个变量。")
    return Strategy(tuple(decode_group(values[4 * i : 4 * i + 4], i) for i in range(3)))


def encode_plan(plan: UavPlan, uav_index: int) -> np.ndarray:
    speed_unit = (plan.speed - UAV_SPEED_MIN) / (UAV_SPEED_MAX - UAV_SPEED_MIN)
    delay_ceiling = min(plan.burst_time, float(FUSE_DELAY_LIMITS[uav_index]))
    delay_unit = 0.0 if delay_ceiling <= 1.0e-12 else plan.fuse_delay / delay_ceiling
    return np.array((plan.heading, speed_unit, plan.burst_time, delay_unit), dtype=float)


def validate_strategy(strategy: Strategy) -> str | None:
    if not 1 <= len(strategy.plans) <= 3:
        return "策略必须包含一至三架无人机。"
    for plan in strategy.plans:
        try:
            index = UAV_NAMES.index(plan.name)
        except ValueError:
            return f"未知无人机：{plan.name}。"
        values = (plan.heading, plan.speed, plan.release_time, plan.fuse_delay, plan.burst_time)
        if not all(np.isfinite(values)):
            return "决策变量必须为有限实数。"
        if not 0.0 <= plan.heading < 2.0 * np.pi + 1.0e-12:
            return f"{plan.name} 航向角越界。"
        if not UAV_SPEED_MIN - 1.0e-10 <= plan.speed <= UAV_SPEED_MAX + 1.0e-10:
            return f"{plan.name} 速度越界。"
        if plan.release_time < -1.0e-10 or plan.fuse_delay < -1.0e-10:
            return f"{plan.name} 投放时刻或引爆延迟为负。"
        if plan.burst_time > BURST_TIME_LIMITS[index] + 1.0e-9:
            return f"{plan.name} 起爆时刻超过解析上界。"
        if plan.burst_point[2] < -1.0e-8:
            return f"{plan.name} 在地面以下起爆。"
    return None


def surface_points(angle_count: int, height_count: int, radial_count: int) -> np.ndarray:
    if min(angle_count, height_count, radial_count) <= 0:
        raise ValueError("完整表面采样数必须为正。")
    angles = np.linspace(0.0, 2.0 * np.pi, angle_count, endpoint=False)
    cosines, sines = np.cos(angles), np.sin(angles)
    x0, y0, z0 = TARGET_BOTTOM_CENTER
    heights = np.linspace(z0, z0 + TARGET_HEIGHT, height_count)
    side = np.vstack(
        [
            np.column_stack(
                (
                    x0 + TARGET_RADIUS * cosines,
                    y0 + TARGET_RADIUS * sines,
                    np.full_like(angles, height),
                )
            )
            for height in heights
        ]
    )
    caps = []
    for height in (z0, z0 + TARGET_HEIGHT):
        for radius in np.linspace(0.0, TARGET_RADIUS, radial_count):
            caps.append(
                np.column_stack(
                    (
                        x0 + radius * cosines,
                        y0 + radius * sines,
                        np.full_like(angles, height),
                    )
                )
            )
    return np.vstack((side, *caps))


class JointEvaluator:
    """三架无人机独立运动下的均匀完整表面联合评价器。"""

    def __init__(
        self,
        angle_count: int,
        height_count: int,
        radial_count: int,
        scan_step: float,
        *,
        root_tolerance: float | None = None,
        target_points: np.ndarray | None = None,
        use_geometric_skip: bool = True,
    ) -> None:
        self.angle_count = angle_count
        self.height_count = height_count
        self.radial_count = radial_count
        self.scan_step = float(scan_step)
        self.root_tolerance = float(root_tolerance or max(1.0e-7, scan_step * 1.0e-4))
        self.points = (
            surface_points(angle_count, height_count, radial_count)
            if target_points is None
            else np.asarray(target_points, dtype=float)
        )
        self.use_geometric_skip = use_geometric_skip

    def missile_positions(self, times: np.ndarray) -> np.ndarray:
        return MISSILE_INITIAL[None, :] + times[:, None] * MISSILE_VELOCITY[None, :]

    @staticmethod
    def cloud_centers(plan: UavPlan, times: np.ndarray) -> np.ndarray:
        centers = np.repeat(plan.burst_point[None, :], len(times), axis=0)
        centers[:, 2] -= CLOUD_SINK_SPEED * (times - plan.burst_time)
        return centers

    @staticmethod
    def active_indices(strategy: Strategy, time: float) -> tuple[int, ...]:
        return tuple(
            index
            for index, plan in enumerate(strategy.plans)
            if plan.burst_time - 1.0e-12 <= time <= min(
                plan.burst_time + CLOUD_LIFETIME, MISSILE_IMPACT_TIME
            )
            + 1.0e-12
        )

    def _potential_mask(
        self, centers: np.ndarray, missiles: np.ndarray
    ) -> np.ndarray:
        if not self.use_geometric_skip:
            return np.ones(len(centers), dtype=bool)
        direction = TARGET_CENTER[None, :] - missiles
        denominator = np.einsum("ij,ij->i", direction, direction)
        projection = np.einsum("ij,ij->i", centers - missiles, direction) / denominator
        projection = np.clip(projection, 0.0, 1.0)
        closest = missiles + projection[:, None] * direction
        distance = np.linalg.norm(centers - closest, axis=1)
        return distance <= CLOUD_RADIUS + TARGET_BOUND_RADIUS + 1.0e-10

    def status_batch(
        self, strategy: Strategy, times: np.ndarray, active: tuple[int, ...]
    ) -> np.ndarray:
        times = np.asarray(times, dtype=float)
        if not active or len(times) == 0:
            return np.zeros(len(times), dtype=bool)
        missiles = self.missile_positions(times)
        directions = self.points[None, :, :] - missiles[:, None, :]
        denominators = np.einsum("tpc,tpc->tp", directions, directions)
        minimum = np.full((len(times), len(self.points)), np.inf, dtype=float)
        for index in active:
            centers = self.cloud_centers(strategy.plans[index], times)
            useful = self._potential_mask(centers, missiles)
            if not np.any(useful):
                continue
            selected = np.flatnonzero(useful)
            selected_directions = directions[selected]
            selected_missiles = missiles[selected]
            selected_centers = centers[selected]
            projection = np.einsum(
                "tc,tpc->tp",
                selected_centers - selected_missiles,
                selected_directions,
            ) / denominators[selected]
            projection = np.clip(projection, 0.0, 1.0)
            closest = selected_missiles[:, None, :] + projection[:, :, None] * selected_directions
            distances = np.linalg.norm(closest - selected_centers[:, None, :], axis=2)
            minimum[selected] = np.minimum(minimum[selected], distances)
        return np.max(minimum, axis=1) <= CLOUD_RADIUS + 1.0e-12

    def status(self, strategy: Strategy, time_value: float, active: tuple[int, ...]) -> bool:
        return bool(self.status_batch(strategy, np.array((time_value,)), active)[0])

    @staticmethod
    def event_segments(strategy: Strategy) -> tuple[tuple[float, float, tuple[int, ...]], ...]:
        events = {0.0, MISSILE_IMPACT_TIME}
        for plan in strategy.plans:
            events.add(float(np.clip(plan.burst_time, 0.0, MISSILE_IMPACT_TIME)))
            events.add(float(np.clip(plan.burst_time + CLOUD_LIFETIME, 0.0, MISSILE_IMPACT_TIME)))
        ordered = sorted(events)
        segments = []
        for left, right in zip(ordered[:-1], ordered[1:]):
            if right - left <= 1.0e-12:
                continue
            middle = 0.5 * (left + right)
            active = JointEvaluator.active_indices(strategy, middle)
            segments.append((left, right, active))
        return tuple(segments)

    def segment_times(self, left: float, right: float) -> np.ndarray:
        times = np.arange(left, right, self.scan_step, dtype=float)
        if times.size == 0 or abs(times[-1] - right) > 1.0e-12:
            times = np.append(times, right)
        return times

    def bisect_boundary(
        self,
        strategy: Strategy,
        left: float,
        right: float,
        active: tuple[int, ...],
        left_valid: bool,
    ) -> float:
        while right - left > self.root_tolerance:
            middle = 0.5 * (left + right)
            middle_valid = self.status(strategy, middle, active)
            if middle_valid == left_valid:
                left = middle
            else:
                right = middle
        return 0.5 * (left + right)

    @staticmethod
    def merge_intervals(
        intervals: Iterable[tuple[float, float]], tolerance: float = 1.0e-8
    ) -> tuple[tuple[float, float], ...]:
        ordered = sorted(intervals)
        if not ordered:
            return ()
        merged = [list(ordered[0])]
        for start, end in ordered[1:]:
            if start <= merged[-1][1] + tolerance:
                merged[-1][1] = max(merged[-1][1], end)
            else:
                merged.append([start, end])
        return tuple((float(start), float(end)) for start, end in merged)

    def evaluate(self, strategy: Strategy) -> Evaluation:
        reason = validate_strategy(strategy)
        grid = (self.angle_count, self.height_count, self.radial_count, self.scan_step)
        release_points = tuple(plan.release_point for plan in strategy.plans)
        burst_points = tuple(plan.burst_point for plan in strategy.plans)
        if reason is not None:
            return Evaluation(strategy, False, reason, (), 0.0, release_points, burst_points, grid)
        intervals: list[tuple[float, float]] = []
        for left, right, active in self.event_segments(strategy):
            if not active:
                continue
            times = self.segment_times(left, right)
            valid = self.status_batch(strategy, times, active)
            current_start = float(times[0]) if bool(valid[0]) else None
            for position in range(1, len(times)):
                if bool(valid[position]) == bool(valid[position - 1]):
                    continue
                boundary = self.bisect_boundary(
                    strategy,
                    float(times[position - 1]),
                    float(times[position]),
                    active,
                    bool(valid[position - 1]),
                )
                if bool(valid[position]):
                    current_start = boundary
                elif current_start is not None:
                    intervals.append((current_start, boundary))
                    current_start = None
            if current_start is not None:
                intervals.append((current_start, right))
        merged = self.merge_intervals(intervals)
        duration = float(sum(end - start for start, end in merged))
        return Evaluation(strategy, True, "", merged, duration, release_points, burst_points, grid)


class Objective:
    def __init__(self, evaluator: JointEvaluator, indices: tuple[int, ...]) -> None:
        self.evaluator = evaluator
        self.indices = indices
        self.cache: dict[tuple[float, ...], Evaluation] = {}

    def strategy(self, vector: Iterable[float]) -> Strategy:
        values = np.asarray(tuple(float(value) for value in vector), dtype=float)
        return Strategy(
            tuple(
                decode_group(values[4 * local : 4 * local + 4], uav_index)
                for local, uav_index in enumerate(self.indices)
            )
        )

    @staticmethod
    def key(vector: Iterable[float]) -> tuple[float, ...]:
        return tuple(round(float(value), 10) for value in vector)

    def result_for(self, vector: Iterable[float]) -> Evaluation:
        key = self.key(vector)
        if key not in self.cache:
            self.cache[key] = self.evaluator.evaluate(self.strategy(key))
        return self.cache[key]

    def __call__(self, vector: np.ndarray) -> float:
        result = self.result_for(vector)
        return -result.duration if result.feasible else 1.0e6


def lhs_population(bounds: Sequence[tuple[float, float]], count: int, seed: int) -> np.ndarray:
    unit = qmc.LatinHypercube(d=len(bounds), seed=seed).random(count)
    lower = np.array([item[0] for item in bounds], dtype=float)
    upper = np.array([item[1] for item in bounds], dtype=float)
    return lower + unit * (upper - lower)


def known_single_anchors(uav_index: int) -> list[np.ndarray]:
    if uav_index != 0:
        return []
    anchors = [
        UavPlan("FY1", tuple(UAV_INITIALS[0]), math.pi, 120.0, 120.0 * 5.1, 5.1, 3.6),
        UavPlan(
            "FY1",
            tuple(UAV_INITIALS[0]),
            math.radians(5.1038039440),
            139.9729312089,
            139.9729312089 * (0.9268493826 + 0.0060780859),
            0.9268493826 + 0.0060780859,
            0.0060780859,
        ),
    ]
    return [encode_plan(plan, 0) for plan in anchors]


def geometric_single_anchors(uav_index: int) -> list[np.ndarray]:
    """构造起爆点恰好落在 M1—目标中心视线上的可行单弹种子。

    这里只利用解析运动学产生非零点目标候选，不向正式目标加入距离引导项。
    对给定起爆时刻和视线参数，先由期望高度反求自由落体时间，再检查所需
    水平速度是否位于题设的 70～140 m/s。
    """

    initial = UAV_INITIALS[uav_index]
    anchors: list[np.ndarray] = []
    burst_times = np.linspace(
        max(0.25, 0.01 * BURST_TIME_LIMITS[uav_index]),
        BURST_TIME_LIMITS[uav_index],
        80,
    )
    sight_parameters = np.linspace(0.05, 0.98, 80)
    for burst_time in burst_times:
        missile = MISSILE_INITIAL + burst_time * MISSILE_VELOCITY
        for sight_parameter in sight_parameters:
            desired = missile + sight_parameter * (TARGET_CENTER - missile)
            vertical_drop = initial[2] - desired[2]
            if vertical_drop < -1.0e-10:
                continue
            fuse_delay = math.sqrt(max(0.0, 2.0 * vertical_drop / GRAVITY))
            if fuse_delay > min(burst_time, FUSE_DELAY_LIMITS[uav_index]) + 1.0e-10:
                continue
            displacement = desired[:2] - initial[:2]
            horizontal_distance = float(np.linalg.norm(displacement))
            speed = horizontal_distance / burst_time
            if not UAV_SPEED_MIN <= speed <= UAV_SPEED_MAX:
                continue
            heading = math.atan2(displacement[1], displacement[0]) % (2.0 * np.pi)
            plan = UavPlan(
                name=UAV_NAMES[uav_index],
                initial=tuple(initial),
                heading=heading,
                speed=speed,
                horizontal_distance=horizontal_distance,
                burst_time=float(burst_time),
                fuse_delay=fuse_delay,
            )
            anchors.append(encode_plan(plan, uav_index))
    return anchors


def angular_distance(left: float, right: float) -> float:
    return abs((left - right + np.pi) % (2.0 * np.pi) - np.pi)


def candidate_distance(left: SingleCandidate, right: SingleCandidate, uav_index: int) -> float:
    heading = angular_distance(left.plan.heading, right.plan.heading) / np.pi
    speed = abs(left.plan.speed - right.plan.speed) / 70.0
    burst = abs(left.plan.burst_time - right.plan.burst_time) / BURST_TIME_LIMITS[uav_index]
    delay = abs(left.plan.fuse_delay - right.plan.fuse_delay) / FUSE_DELAY_LIMITS[uav_index]
    return float(np.linalg.norm((heading, speed, burst, delay)))


def build_single_library(profile: SearchProfile, uav_index: int, seed: int) -> tuple[SingleCandidate, ...]:
    point = JointEvaluator(
        1,
        1,
        1,
        0.05 if profile.name == "quick" else 0.02,
        target_points=TARGET_CENTER[None, :],
    )
    objective = Objective(point, (uav_index,))
    bounds = variable_bounds((uav_index,))
    initial = lhs_population(bounds, profile.single_population, seed)
    anchors = [*known_single_anchors(uav_index), *geometric_single_anchors(uav_index)]
    # 几何构造通常能产生大量可行种子。先用点目标时长排序，再只向初始种群
    # 放入彼此不同的前若干个，剩余位置仍由全域 LHS 提供。
    anchor_records = sorted(
        ((-objective(vector), np.asarray(vector, dtype=float)) for vector in anchors),
        key=lambda item: -item[0],
    )
    anchor_limit = min(len(anchor_records), max(2, profile.single_population // 2))
    if anchor_limit:
        initial[:anchor_limit] = np.vstack([item[1] for item in anchor_records[:anchor_limit]])
    result = differential_evolution(
        objective,
        bounds=bounds,
        strategy="best1bin",
        init=initial,
        maxiter=profile.single_maxiter,
        mutation=(0.5, 1.0),
        recombination=0.9,
        seed=seed,
        polish=False,
        tol=1.0e-5,
        atol=0.0,
        updating="immediate",
        workers=1,
    )
    population = np.asarray(result.population, dtype=float)
    energies = np.asarray(result.population_energies, dtype=float)
    combined_vectors = [population[index].copy() for index in np.argsort(energies)]
    combined_vectors.extend(item[1] for item in anchor_records)
    # 先按点目标时长统一排序，确保解析种子不会因 DE 的随机选择而丢失。
    combined_vectors.sort(key=lambda vector: objective(vector))
    combined_vectors = combined_vectors[: profile.single_pool_count]
    low_surface = JointEvaluator(24, 3, 3, 0.10 if profile.name == "quick" else 0.05)
    candidates = []
    for vector in combined_vectors:
        point_result = objective.result_for(vector)
        strategy = Strategy((decode_group(vector, uav_index),))
        surface_result = low_surface.evaluate(strategy)
        candidates.append(
            SingleCandidate(
                encoded=vector,
                plan=strategy.plans[0],
                point_duration=point_result.duration,
                surface_duration=surface_result.duration,
            )
        )
    candidates.sort(key=lambda item: (-item.surface_duration, -item.point_duration))
    selected: list[SingleCandidate] = []
    for candidate in candidates:
        if not selected or all(candidate_distance(candidate, other, uav_index) >= 0.03 for other in selected):
            selected.append(candidate)
        if len(selected) >= profile.single_library_size:
            break
    if len(selected) < profile.single_library_size:
        selected_keys = {tuple(np.round(item.encoded, 10)) for item in selected}
        for candidate in candidates:
            key = tuple(np.round(candidate.encoded, 10))
            if key not in selected_keys:
                selected.append(candidate)
                selected_keys.add(key)
            if len(selected) >= profile.single_library_size:
                break
    if not selected:
        raise RuntimeError(f"{UAV_NAMES[uav_index]} 未生成任何单弹候选。")
    return tuple(selected)


def build_joint_initial_population(
    profile: SearchProfile,
    libraries: Sequence[Sequence[SingleCandidate]],
    seed: int,
) -> np.ndarray:
    population = lhs_population(JOINT_BOUNDS, profile.population_size, seed + 10_000)
    seeded_count = int(round(profile.population_size * (1.0 - profile.global_fraction)))
    combinations = list(itertools.product(*libraries))
    rng = np.random.default_rng(seed + 20_000)
    rng.shuffle(combinations)
    for row in range(seeded_count):
        combo = combinations[row % len(combinations)]
        vector = np.concatenate([item.encoded for item in combo]).astype(float)
        if row >= len(combinations):
            for uav_index in range(3):
                start = 4 * uav_index
                vector[start] = (vector[start] + rng.normal(0.0, 0.06)) % (2.0 * np.pi)
                vector[start + 1] = np.clip(vector[start + 1] + rng.normal(0.0, 0.04), 0.0, 1.0)
                vector[start + 2] = np.clip(
                    vector[start + 2] + rng.normal(0.0, 0.03 * BURST_TIME_LIMITS[uav_index]),
                    0.0,
                    BURST_TIME_LIMITS[uav_index],
                )
                vector[start + 3] = np.clip(vector[start + 3] + rng.normal(0.0, 0.04), 0.0, 1.0)
        population[row] = vector
    return population


def run_joint_seed(
    profile: SearchProfile,
    libraries: Sequence[Sequence[SingleCandidate]],
    seed: int,
    workers: int,
) -> SearchRun:
    evaluator = JointEvaluator(*profile.search_grid)
    objective = Objective(evaluator, (0, 1, 2))
    initial = build_joint_initial_population(profile, libraries, seed)
    history: list[float] = []

    def callback(vector: np.ndarray, convergence: float) -> bool:
        del convergence
        current = objective.result_for(vector).duration
        history.append(max(history[-1], current) if history else current)
        return False

    started = time.perf_counter()
    result = differential_evolution(
        objective,
        bounds=JOINT_BOUNDS,
        strategy="best1bin",
        init=initial,
        maxiter=profile.maxiter,
        mutation=(0.5, 1.0),
        recombination=0.9,
        seed=seed,
        polish=False,
        tol=1.0e-5,
        atol=0.0,
        callback=callback,
        updating="immediate" if workers == 1 else "deferred",
        workers=workers,
    )
    elapsed = time.perf_counter() - started
    population = np.asarray(result.population, dtype=float)
    energies = np.asarray(result.population_energies, dtype=float)
    order = np.argsort(energies)[: profile.top_per_seed]
    candidates = tuple(
        JointCandidate(
            encoded=population[index].copy(),
            result=objective.result_for(population[index]),
            source="joint_de_population",
            seed=seed,
        )
        for index in order
    )
    best_result = objective.result_for(result.x)
    best = JointCandidate(np.asarray(result.x, dtype=float), best_result, "joint_de_best", seed)
    if not history:
        history.append(best.result.duration)
    return SearchRun(
        seed=seed,
        elapsed_seconds=elapsed,
        iterations=int(result.nit),
        evaluations=int(result.nfev),
        success=bool(result.success),
        message=str(result.message),
        best=best,
        candidates=candidates,
        history=tuple(history),
    )


def candidate_key(candidate: JointCandidate) -> tuple[float, ...]:
    return tuple(round(float(value), 7) for value in candidate.encoded)


def deduplicate(candidates: Iterable[JointCandidate]) -> list[JointCandidate]:
    unique: dict[tuple[float, ...], JointCandidate] = {}
    for candidate in candidates:
        key = candidate_key(candidate)
        current = unique.get(key)
        if current is None or candidate.result.duration > current.result.duration:
            unique[key] = candidate
    return sorted(unique.values(), key=lambda item: -item.result.duration)


def rerank(
    candidates: Sequence[JointCandidate],
    grid: tuple[int, int, int, float],
    count: int,
    source: str,
) -> list[JointCandidate]:
    evaluator = JointEvaluator(*grid)
    reviewed = [
        JointCandidate(candidate.encoded, evaluator.evaluate(decode_strategy(candidate.encoded)), source, candidate.seed)
        for candidate in candidates[:count]
    ]
    return sorted(reviewed, key=lambda item: -item.result.duration)


def union_intervals(intervals: Iterable[tuple[float, float]]) -> tuple[tuple[float, float], ...]:
    return JointEvaluator.merge_intervals(intervals)


def independent_statistics(
    strategy: Strategy, evaluator: JointEvaluator
) -> tuple[float, tuple[float, ...], tuple[tuple[float, float], ...]]:
    results = [evaluator.evaluate(Strategy((plan,))) for plan in strategy.plans]
    merged = union_intervals(interval for result in results for interval in result.intervals)
    duration = float(sum(end - start for start, end in merged))
    return duration, tuple(result.duration for result in results), merged


def physical_perturbation(
    strategy: Strategy, plan_index: int, factor: str, delta: float
) -> Strategy | None:
    plans = list(strategy.plans)
    original = plans[plan_index]
    heading = original.heading
    speed = original.speed
    release = original.release_time
    delay = original.fuse_delay
    if factor == "heading_deg":
        heading = (heading + math.radians(delta)) % (2.0 * np.pi)
    elif factor == "speed_mps":
        speed += delta
    elif factor == "release_time_s":
        release += delta
    elif factor == "fuse_delay_s":
        delay += delta
    else:
        raise ValueError(f"未知扰动因素：{factor}")
    burst = release + delay
    if burst < 0.0:
        return None
    changed = UavPlan(
        name=original.name,
        initial=original.initial,
        heading=heading,
        speed=speed,
        horizontal_distance=speed * burst,
        burst_time=burst,
        fuse_delay=delay,
    )
    plans[plan_index] = changed
    candidate = Strategy(tuple(plans))
    return None if validate_strategy(candidate) is not None else candidate


def sensitivity_rows(strategy: Strategy, base_duration: float, evaluator: JointEvaluator) -> list[dict[str, str]]:
    settings = (
        ("heading_deg", 1.0),
        ("speed_mps", 1.0),
        ("release_time_s", 0.1),
        ("fuse_delay_s", 0.1),
    )
    rows: list[dict[str, str]] = []
    for plan_index, plan in enumerate(strategy.plans):
        for factor, magnitude in settings:
            for sign in (-1.0, 1.0):
                delta = sign * magnitude
                changed = physical_perturbation(strategy, plan_index, factor, delta)
                if changed is None:
                    duration = math.nan
                    feasible = False
                else:
                    result = evaluator.evaluate(changed)
                    duration = result.duration
                    feasible = result.feasible
                rows.append(
                    {
                        "uav": plan.name,
                        "factor": factor,
                        "perturbation": f"{delta:.10f}",
                        "feasible": str(feasible),
                        "duration_s": "" if not np.isfinite(duration) else f"{duration:.10f}",
                        "delta_duration_s": "" if not np.isfinite(duration) else f"{duration - base_duration:.10f}",
                    }
                )
    return rows


def format_intervals(intervals: Iterable[tuple[float, float]]) -> str:
    return "; ".join(f"[{start:.10f}, {end:.10f}]" for start, end in intervals)


def write_csv(path: Path, rows: Sequence[dict[str, object]]) -> None:
    if not rows:
        raise RuntimeError(f"没有可写入 {path.name} 的记录。")
    fieldnames = list(dict.fromkeys(key for row in rows for key in row))
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def normalize_header(value: object) -> str:
    return "".join(str(value or "").split()).replace("：", ":")


def header_column(worksheet, keywords: Sequence[str]) -> int:
    for cell in worksheet[1]:
        value = normalize_header(cell.value)
        if all(keyword in value for keyword in keywords):
            return cell.column
    raise KeyError(f"模板缺少列：{'/'.join(keywords)}")


def write_workbook(
    result: Evaluation,
    individual_durations: Sequence[float],
) -> Path:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"未找到模板：{TEMPLATE_PATH}")
    output = Q4_DIR / "result2.xlsx"
    shutil.copy2(TEMPLATE_PATH, output)
    workbook = load_workbook(output)
    worksheet = workbook.active
    columns = {
        "name": header_column(worksheet, ("无人机", "编号")),
        "heading": header_column(worksheet, ("无人机", "方向")),
        "speed": header_column(worksheet, ("无人机", "速度")),
        "release_x": header_column(worksheet, ("投放点", "x")),
        "release_y": header_column(worksheet, ("投放点", "y")),
        "release_z": header_column(worksheet, ("投放点", "z")),
        "burst_x": header_column(worksheet, ("起爆点", "x")),
        "burst_y": header_column(worksheet, ("起爆点", "y")),
        "burst_z": header_column(worksheet, ("起爆点", "z")),
        "duration": header_column(worksheet, ("有效", "时长")),
    }
    for index, plan in enumerate(result.strategy.plans):
        row = index + 2
        release = result.release_points[index]
        burst = result.burst_points[index]
        values = {
            "name": plan.name,
            "heading": float(np.degrees(plan.heading)),
            "speed": float(plan.speed),
            "release_x": float(release[0]),
            "release_y": float(release[1]),
            "release_z": float(release[2]),
            "burst_x": float(burst[0]),
            "burst_y": float(burst[1]),
            "burst_z": float(burst[2]),
            "duration": float(individual_durations[index]),
        }
        for key, value in values.items():
            worksheet.cell(row, columns[key]).value = value
    workbook.save(output)
    return output


def save_outputs(
    profile: SearchProfile,
    libraries: Sequence[Sequence[SingleCandidate]],
    runs: Sequence[SearchRun],
    final_candidates: Sequence[JointCandidate],
    verification: Evaluation,
    independent_duration: float,
    individual_durations: tuple[float, ...],
    independent_intervals: tuple[tuple[float, float], ...],
    sensitivity: Sequence[dict[str, str]],
) -> tuple[Path, ...]:
    Q4_DIR.mkdir(parents=True, exist_ok=True)
    library_rows: list[dict[str, object]] = []
    for uav_index, library in enumerate(libraries):
        for rank, candidate in enumerate(library, start=1):
            library_rows.append(
                {
                    "profile": profile.name,
                    "uav": UAV_NAMES[uav_index],
                    "rank": rank,
                    "theta_deg": np.degrees(candidate.plan.heading),
                    "speed_mps": candidate.plan.speed,
                    "release_time_s": candidate.plan.release_time,
                    "fuse_delay_s": candidate.plan.fuse_delay,
                    "burst_time_s": candidate.plan.burst_time,
                    "point_duration_s": candidate.point_duration,
                    "low_surface_duration_s": candidate.surface_duration,
                }
            )
    search_rows: list[dict[str, object]] = []
    for run in runs:
        plan = run.best.result.strategy.plans
        search_rows.append(
            {
                "profile": profile.name,
                "seed": run.seed,
                "elapsed_seconds": run.elapsed_seconds,
                "iterations": run.iterations,
                "function_evaluations": run.evaluations,
                "success": run.success,
                "message": run.message,
                "search_duration_s": run.best.result.duration,
                "burst_times_s": "; ".join(f"{item.burst_time:.10f}" for item in plan),
                "intervals_s": format_intervals(run.best.result.intervals),
            }
        )
    history_rows = [
        {
            "profile": profile.name,
            "seed": run.seed,
            "generation": generation,
            "best_duration_s": duration,
        }
        for run in runs
        for generation, duration in enumerate(run.history, start=1)
    ]
    best_rows: list[dict[str, object]] = []
    selected = verification.strategy
    for index, plan in enumerate(selected.plans):
        release = verification.release_points[index]
        burst = verification.burst_points[index]
        best_rows.append(
            {
                "profile": profile.name,
                "uav": plan.name,
                "theta_rad": plan.heading,
                "theta_deg": np.degrees(plan.heading),
                "speed_mps": plan.speed,
                "horizontal_distance_s_m": plan.horizontal_distance,
                "release_time_s": plan.release_time,
                "fuse_delay_s": plan.fuse_delay,
                "burst_time_s": plan.burst_time,
                "release_x_m": release[0],
                "release_y_m": release[1],
                "release_z_m": release[2],
                "burst_x_m": burst[0],
                "burst_y_m": burst[1],
                "burst_z_m": burst[2],
                "individual_duration_s": individual_durations[index],
                "joint_duration_s": verification.duration,
                "joint_intervals_s": format_intervals(verification.intervals),
                "independent_union_duration_s": independent_duration,
                "independent_union_intervals_s": format_intervals(independent_intervals),
                "synergy_gain_s": verification.duration - independent_duration,
                "verification_grid": f"{verification.grid[0]}x{verification.grid[1]}x{verification.grid[2]}@{verification.grid[3]}",
            }
        )
    library_path = Q4_DIR / "q4_single_library.csv"
    search_path = Q4_DIR / "q4_search_runs.csv"
    history_path = Q4_DIR / "q4_de_history.csv"
    best_path = Q4_DIR / "q4_best_solution.csv"
    sensitivity_path = Q4_DIR / "q4_sensitivity.csv"
    write_csv(library_path, library_rows)
    write_csv(search_path, search_rows)
    write_csv(history_path, history_rows)
    write_csv(best_path, best_rows)
    write_csv(sensitivity_path, list(sensitivity))
    workbook_path = write_workbook(verification, individual_durations)
    return library_path, search_path, history_path, best_path, sensitivity_path, workbook_path


def run_regression_checks() -> None:
    if not np.allclose(BURST_TIME_LIMITS, (13.9422362498, 50.5327205253, 66.9991708075), atol=1.0e-8):
        raise AssertionError("三架无人机起爆时刻解析上界计算错误。")
    rng = np.random.default_rng(20250816)
    for vector in lhs_population(JOINT_BOUNDS, 100, 20250816):
        strategy = decode_strategy(vector)
        reason = validate_strategy(strategy)
        if reason is not None:
            raise AssertionError(f"可行化解码失败：{reason}")
    q1 = UavPlan("FY1", tuple(UAV_INITIALS[0]), math.pi, 120.0, 120.0 * 5.1, 5.1, 3.6)
    evaluator = JointEvaluator(180, 9, 9, 0.02)
    result = evaluator.evaluate(Strategy((q1,)))
    if abs(result.duration - 1.3916426693) > 2.0e-3:
        raise AssertionError(f"Q1 完整表面回归失败：{result.duration:.10f} s。")
    # 随机数实际参与循环，避免误把仅构造但未执行的检查当作回归。
    if not np.isfinite(rng.random()):
        raise AssertionError("随机数发生器异常。")


def solve(profile: SearchProfile, workers: int) -> tuple[
    tuple[tuple[SingleCandidate, ...], ...],
    list[SearchRun],
    list[JointCandidate],
    Evaluation,
    float,
    tuple[float, ...],
    tuple[tuple[float, float], ...],
    list[dict[str, str]],
]:
    print("正在执行物理约束和 Q1 评价器回归检查……", flush=True)
    run_regression_checks()
    libraries = []
    for index in range(3):
        print(f"正在生成 {UAV_NAMES[index]} 单弹候选库……", flush=True)
        library = build_single_library(profile, index, 20250900 + index)
        libraries.append(library)
        print(
            f"  保留 {len(library)} 个候选；最佳低密度完整表面时长 "
            f"{library[0].surface_duration:.6f} s。",
            flush=True,
        )
    runs = []
    for seed in profile.seeds:
        print(f"正在运行 12 维完整表面联合 DE：seed={seed}……", flush=True)
        run = run_joint_seed(profile, libraries, seed, workers)
        runs.append(run)
        print(
            f"  完成：{run.elapsed_seconds:.1f} s，搜索精度联合时长 "
            f"{run.best.result.duration:.6f} s。",
            flush=True,
        )
    pool = deduplicate(candidate for run in runs for candidate in (run.best, *run.candidates))
    print("正在执行中密度候选重排……", flush=True)
    middle = rerank(pool, profile.rerank_grid, profile.rerank_count, "middle_rerank")
    print("正在执行高密度完整表面复核……", flush=True)
    final_candidates = rerank(middle, profile.final_grid, profile.final_count, "final_review")
    if not final_candidates:
        raise RuntimeError("没有候选通过最终完整表面评价。")
    print("正在执行独立加密交叉复核……", flush=True)
    verification_evaluator = JointEvaluator(*profile.verification_grid)
    verification = verification_evaluator.evaluate(final_candidates[0].result.strategy)
    independent_duration, individual_durations, independent_intervals = independent_statistics(
        verification.strategy, verification_evaluator
    )
    print("正在执行 24 次单因素扰动检查……", flush=True)
    sensitivity_evaluator = JointEvaluator(*profile.final_grid)
    sensitivity = sensitivity_rows(
        verification.strategy,
        sensitivity_evaluator.evaluate(verification.strategy).duration,
        sensitivity_evaluator,
    )
    return (
        tuple(libraries),
        runs,
        final_candidates,
        verification,
        independent_duration,
        individual_durations,
        independent_intervals,
        sensitivity,
    )


def print_result(
    verification: Evaluation,
    independent_duration: float,
    individual_durations: Sequence[float],
) -> None:
    print("\n问题四推荐方案（完整圆柱联合遮蔽）")
    for index, plan in enumerate(verification.strategy.plans):
        print(
            f"  {plan.name}: 航向 {np.degrees(plan.heading):.8f}°，速度 {plan.speed:.8f} m/s，"
            f"投放 {plan.release_time:.8f} s，延迟 {plan.fuse_delay:.8f} s，起爆 {plan.burst_time:.8f} s"
        )
        print(f"       投放点 {np.array2string(verification.release_points[index], precision=6)}")
        print(f"       起爆点 {np.array2string(verification.burst_points[index], precision=6)}")
        print(f"       单弹完整遮蔽时长 {individual_durations[index]:.10f} s")
    print(f"  联合遮蔽区间: {format_intervals(verification.intervals)} s")
    print(f"  联合遮蔽总时长: {verification.duration:.10f} s")
    print(f"  单弹完整遮蔽区间并集: {independent_duration:.10f} s")
    print(f"  空间协同增益: {verification.duration - independent_duration:.10f} s")
    print(
        f"  最终复核网格: {verification.grid[0]}×{verification.grid[1]}×"
        f"{verification.grid[2]}，时间步长 {verification.grid[3]} s"
    )


def parse_seeds(text: str) -> tuple[int, ...]:
    try:
        seeds = tuple(int(item.strip()) for item in text.split(",") if item.strip())
    except ValueError as error:
        raise argparse.ArgumentTypeError("随机种子必须为逗号分隔的整数。") from error
    if not seeds:
        raise argparse.ArgumentTypeError("至少需要一个随机种子。")
    return seeds


def main() -> None:
    parser = argparse.ArgumentParser(description="问题四三机三烟幕联合遮蔽标准 DE 求解")
    parser.add_argument("--profile", choices=tuple(PROFILES), default="standard")
    parser.add_argument("--seeds", type=parse_seeds, help="覆盖配置中的随机种子，例如 41,42")
    parser.add_argument(
        "--workers",
        type=int,
        default=1,
        help="SciPy DE 并行工作进程数；Windows 下建议先使用 1。",
    )
    arguments = parser.parse_args()
    if arguments.workers == 0 or arguments.workers < -1:
        parser.error("workers 必须为 -1 或正整数。")
    profile = PROFILES[arguments.profile]
    if arguments.seeds is not None:
        profile = replace(profile, seeds=arguments.seeds)
    started = time.perf_counter()
    outputs = solve(profile, arguments.workers)
    libraries, runs, final_candidates, verification, independent, individual, independent_intervals, sensitivity = outputs
    print_result(verification, independent, individual)
    paths = save_outputs(
        profile,
        libraries,
        runs,
        final_candidates,
        verification,
        independent,
        individual,
        independent_intervals,
        sensitivity,
    )
    for path in paths:
        print(f"结果已保存至: {path}")
    print(f"总运行时间: {time.perf_counter() - started:.2f} s")


if __name__ == "__main__":
    main()

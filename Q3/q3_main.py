"""问题三：三枚烟幕联合完整遮蔽的标准差分进化求解。"""

from __future__ import annotations

import argparse
import csv
import math
import shutil
import sys
import time
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Iterable, Sequence

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import differential_evolution
from scipy.stats import qmc

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core.multi_smoke_evaluator import (  # noqa: E402
    AdaptiveSurfaceConfig,
    JointEvaluationResult,
    MultiSmokeEvaluator,
    ThreeDeployment,
    ThreeSmokeSimulation,
    TimeDiagnostic,
    UniformReview,
)
from core.smoke_evaluator import ScenarioParameters  # noqa: E402

Q3_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = PROJECT_ROOT / "Resources" / "result1.xlsx"
MAX_BURST_TIME = 13.9423
DIMENSION = 8
BOUNDS = ((0.0, 2.0 * np.pi), (70.0, 140.0), *((0.0, 1.0),) * 6)
QUICK_SEEDS = (1, 2)
STANDARD_SEEDS = (1, 2, 3)


@dataclass(frozen=True)
class SearchProfile:
    name: str
    seeds: tuple[int, ...]
    point_population_size: int
    point_maxiter: int
    population_size: int
    maxiter: int
    search_grid: tuple[int, int, int, float]
    rerank_grid: tuple[int, int, int, float]
    final_grid: tuple[int, int, int, float]
    verification_grid: tuple[int, int, int, float]
    adaptive: AdaptiveSurfaceConfig
    rerank_adaptive: AdaptiveSurfaceConfig
    final_adaptive: AdaptiveSurfaceConfig
    top_per_seed: int = 10
    rerank_count: int = 10
    final_count: int = 3


PROFILES = {
    "quick": SearchProfile(
        name="quick",
        seeds=QUICK_SEEDS,
        point_population_size=80,
        point_maxiter=150,
        population_size=32,
        maxiter=35,
        search_grid=(24, 3, 3, 0.15),
        rerank_grid=(60, 5, 5, 0.05),
        final_grid=(180, 9, 9, 0.02),
        verification_grid=(360, 11, 11, 0.01),
        adaptive=AdaptiveSurfaceConfig(
            rho_min=1.0, max_depth=6, max_patches=5_000, scan_step=0.20, root_tolerance=1.0e-4
        ),
        rerank_adaptive=AdaptiveSurfaceConfig(
            rho_min=0.5, max_depth=8, max_patches=20_000, scan_step=0.10, root_tolerance=1.0e-5
        ),
        final_adaptive=AdaptiveSurfaceConfig(
            rho_min=0.25, max_depth=10, max_patches=50_000, scan_step=0.05, root_tolerance=1.0e-6
        ),
        rerank_count=6,
        final_count=2,
    ),
    "standard": SearchProfile(
        name="standard",
        seeds=STANDARD_SEEDS,
        point_population_size=80,
        point_maxiter=180,
        # 点目标阶段负责全域找方向；完整表面阶段承担精调，因此将其
        # 搜索强度提高到 64 个体、120 代，同时保持三种子可并行。
        population_size=64,
        maxiter=120,
        search_grid=(36, 5, 5, 0.10),
        rerank_grid=(120, 7, 7, 0.02),
        final_grid=(360, 11, 11, 0.01),
        verification_grid=(720, 21, 15, 0.005),
        adaptive=AdaptiveSurfaceConfig(
            rho_min=0.5, max_depth=8, max_patches=20_000, scan_step=0.10, root_tolerance=1.0e-5
        ),
        rerank_adaptive=AdaptiveSurfaceConfig(
            rho_min=0.25, max_depth=10, max_patches=50_000, scan_step=0.05, root_tolerance=1.0e-6
        ),
        final_adaptive=AdaptiveSurfaceConfig(
            rho_min=0.10, max_depth=12, max_patches=100_000, scan_step=0.02, root_tolerance=1.0e-6
        ),
        # 低密度网格下的排名可能变化。扩大传递给高精度评价器的候选池，
        # 比只在同一低密度网格上盲目增加迭代更能降低误排序风险。
        top_per_seed=20,
        rerank_count=20,
        final_count=5,
    ),
}


@dataclass(frozen=True)
class Candidate:
    vector: np.ndarray
    deployment: ThreeDeployment
    result: JointEvaluationResult
    source: str
    seed: int | str
    rank: int


@dataclass(frozen=True)
class SearchRun:
    seed: int
    elapsed_seconds: float
    iterations: int
    evaluations: int
    success: bool
    message: str
    best: Candidate
    candidates: tuple[Candidate, ...]
    history: tuple[float, ...]


@dataclass(frozen=True)
class PointSearchResult:
    vector: np.ndarray
    population: np.ndarray
    energies: np.ndarray
    evaluations: int
    message: str


@dataclass(frozen=True)
class FinalReview:
    candidate: Candidate
    result: JointEvaluationResult
    independent_duration: float
    individual_durations: tuple[float, ...]
    uniform_review: UniformReview | None = None
    dense_uniform_review: UniformReview | None = None
    adaptive_review: JointEvaluationResult | None = None


def decode_vector(vector: Iterable[float]) -> ThreeDeployment:
    """将 DE 单位区间变量还原为严格可行的物理策略。"""

    values = tuple(float(value) for value in vector)
    if len(values) != DIMENSION:
        raise ValueError("问题三 DE 决策向量必须包含八个变量。")
    heading, speed, u1, u2, u3, u4, u5, u6 = values
    u1, u2, u3, u4, u5, u6 = (float(np.clip(value, 0.0, 1.0)) for value in (u1, u2, u3, u4, u5, u6))
    release1 = (MAX_BURST_TIME - 2.0) * u1
    slack = MAX_BURST_TIME - release1 - 2.0
    gap12 = 1.0 + slack * u2
    gap23 = 1.0 + slack * (1.0 - u2) * u3
    releases = (release1, release1 + gap12, release1 + gap12 + gap23)
    delays = tuple((MAX_BURST_TIME - release) * unit for release, unit in zip(releases, (u4, u5, u6)))
    return ThreeDeployment(
        heading=float(np.clip(heading, 0.0, 2.0 * np.pi)),
        speed=float(np.clip(speed, 70.0, 140.0)),
        release_times=tuple(float(value) for value in releases),
        fuse_delays=tuple(float(value) for value in delays),
    )


def encode_deployment(deployment: ThreeDeployment) -> np.ndarray:
    """将物理策略编码回单位区间，用于已知可行种子。"""

    if len(deployment.release_times) != 3:
        raise ValueError("问题三种子必须包含三枚烟幕弹。")
    release1, release2, release3 = deployment.release_times
    gap12, gap23 = release2 - release1, release3 - release2
    slack = MAX_BURST_TIME - release1 - 2.0
    if slack < -1.0e-10 or min(gap12, gap23) < 1.0 - 1.0e-10:
        raise ValueError("种子方案不满足问题三投放间隔。")
    u1 = release1 / (MAX_BURST_TIME - 2.0)
    u2 = 0.0 if slack <= 1.0e-12 else (gap12 - 1.0) / slack
    remaining = slack - (gap12 - 1.0)
    u3 = 0.0 if remaining <= 1.0e-12 else (gap23 - 1.0) / remaining
    units = [u1, u2, u3]
    for release, delay in zip(deployment.release_times, deployment.fuse_delays):
        units.append(0.0 if MAX_BURST_TIME - release <= 1.0e-12 else delay / (MAX_BURST_TIME - release))
    return np.array(
        [deployment.heading, deployment.speed, *np.clip(units, 0.0, 1.0)], dtype=float
    )


def is_feasible_deployment(deployment: ThreeDeployment) -> bool:
    if len(deployment.release_times) != 3:
        return False
    releases = deployment.release_times
    return (
        0.0 <= deployment.heading <= 2.0 * np.pi
        and 70.0 <= deployment.speed <= 140.0
        and releases[1] - releases[0] >= 1.0 - 1.0e-10
        and releases[2] - releases[1] >= 1.0 - 1.0e-10
        and all(release >= 0.0 and delay >= 0.0 for release, delay in zip(releases, deployment.fuse_delays))
        and all(burst <= MAX_BURST_TIME + 1.0e-10 for burst in deployment.burst_times)
    )


def build_evaluator(config: AdaptiveSurfaceConfig) -> MultiSmokeEvaluator:
    return MultiSmokeEvaluator(
        ScenarioParameters(max_burst_time=MAX_BURST_TIME), config
    )


class UniformJointEvaluator:
    """缓存完整圆柱表面点的快速联合评价器。

    它与自适应评价器使用完全相同的运动学、有效烟幕集合和
    ``max_Q min_j d_j`` 联合判据，只把连续表面换成可逐级加密的均匀网格。
    """

    def __init__(
        self,
        angle_count: int,
        height_count: int,
        radial_count: int,
        scan_step: float,
        root_tolerance: float,
    ) -> None:
        self.angle_count = angle_count
        self.height_count = height_count
        self.radial_count = radial_count
        config = AdaptiveSurfaceConfig(
            scan_step=scan_step,
            root_tolerance=root_tolerance,
        )
        self.evaluator = build_evaluator(config)
        self.points = ThreeSmokeSimulation.uniform_surface_points(
            self.evaluator.parameters,
            angle_count,
            height_count,
            radial_count,
        )

    def evaluate(self, deployment: ThreeDeployment) -> JointEvaluationResult:
        reason = self.evaluator.validate(deployment)
        if reason is not None:
            return JointEvaluationResult(
                deployment=deployment,
                feasible=False,
                reason=reason,
                release_points=(),
                burst_points=(),
                intervals=(),
                duration=0.0,
            )
        simulation = self.evaluator.simulation(deployment)
        assert simulation is not None
        review = simulation.uniform_review_points(
            self.points,
            angle_count=self.angle_count,
            height_count=self.height_count,
            radial_count=self.radial_count,
        )
        return JointEvaluationResult(
            deployment=deployment,
            feasible=True,
            reason="",
            release_points=tuple(point.copy() for point in simulation.release_points),
            burst_points=tuple(point.copy() for point in simulation.burst_points),
            intervals=review.intervals,
            duration=review.duration,
        )

    def review_from_result(self, result: JointEvaluationResult) -> UniformReview:
        return UniformReview(
            intervals=result.intervals,
            duration=result.duration,
            angle_count=self.angle_count,
            height_count=self.height_count,
            radial_count=self.radial_count,
        )


class PointTargetObjective:
    """只用于产生全域初值的圆柱中心点近似目标。"""

    def __init__(self, step: float = 0.05) -> None:
        self.step = step
        self.parameters = ScenarioParameters(max_burst_time=MAX_BURST_TIME)
        self.times = np.arange(
            0.0,
            MAX_BURST_TIME + self.parameters.cloud_lifetime + 0.5 * step,
            step,
        )
        missile_initial = np.asarray(self.parameters.missile_initial, dtype=float)
        missile_velocity = (
            -self.parameters.missile_speed
            * missile_initial
            / np.linalg.norm(missile_initial)
        )
        self.missiles = missile_initial[None, :] + self.times[:, None] * missile_velocity
        x0, y0, z0 = self.parameters.target_bottom_center
        self.target = np.array(
            [x0, y0, z0 + 0.5 * self.parameters.target_height], dtype=float
        )
        self.directions = self.target[None, :] - self.missiles
        self.denominators = np.einsum(
            "ij,ij->i", self.directions, self.directions
        )
        self.uav_initial = np.asarray(self.parameters.uav_initial, dtype=float)
        self.cache: dict[tuple[float, ...], tuple[float, float]] = {}

    @staticmethod
    def key(vector: Iterable[float]) -> tuple[float, ...]:
        return tuple(round(float(value), 12) for value in vector)

    def metrics(self, vector: Iterable[float]) -> tuple[float, float]:
        key = self.key(vector)
        if key in self.cache:
            return self.cache[key]
        deployment = decode_vector(key)
        direction = np.array(
            [np.cos(deployment.heading), np.sin(deployment.heading), 0.0],
            dtype=float,
        )
        covered = np.zeros(len(self.times), dtype=bool)
        guide = 0.0
        for release, delay, burst in zip(
            deployment.release_times,
            deployment.fuse_delays,
            deployment.burst_times,
        ):
            burst_point = np.array(
                [
                    self.uav_initial[0]
                    + deployment.speed * burst * direction[0],
                    self.uav_initial[1]
                    + deployment.speed * burst * direction[1],
                    self.uav_initial[2]
                    - 0.5 * self.parameters.gravity * delay**2,
                ],
                dtype=float,
            )
            centers = np.repeat(burst_point[None, :], len(self.times), axis=0)
            centers[:, 2] -= self.parameters.cloud_sink_speed * (
                self.times - burst
            )
            center_from_missile = centers - self.missiles
            projection = np.einsum(
                "ij,ij->i", center_from_missile, self.directions
            ) / self.denominators
            projection = np.clip(projection, 0.0, 1.0)
            closest = self.missiles + projection[:, None] * self.directions
            distances = np.linalg.norm(centers - closest, axis=1)
            active = (self.times >= burst) & (
                self.times <= burst + self.parameters.cloud_lifetime
            )
            covered |= active & (distances <= self.parameters.cloud_radius)
            guide += float(np.min(distances[active]))
        duration = float(np.sum(covered) * self.step)
        self.cache[key] = (duration, guide)
        return duration, guide

    def __call__(self, vector: np.ndarray) -> float:
        duration, guide = self.metrics(vector)
        # 时长始终是主目标；极小的几何距离项只用于打破相同离散时长的平台。
        return -duration + 1.0e-4 * guide


class DurationObjective:
    def __init__(
        self, evaluator: MultiSmokeEvaluator | UniformJointEvaluator
    ) -> None:
        self.evaluator = evaluator
        self.cache: dict[tuple[float, ...], JointEvaluationResult] = {}

    @staticmethod
    def key(vector: Iterable[float]) -> tuple[float, ...]:
        return tuple(round(float(value), 12) for value in vector)

    def result_for(self, vector: Iterable[float]) -> JointEvaluationResult:
        key = self.key(vector)
        if key not in self.cache:
            self.cache[key] = self.evaluator.evaluate(decode_vector(key))
        return self.cache[key]

    def __call__(self, vector: np.ndarray) -> float:
        result = self.result_for(vector)
        return -result.duration if result.feasible else 1.0e6


def q1_anchor() -> ThreeDeployment:
    return ThreeDeployment(
        heading=math.pi,
        speed=120.0,
        release_times=(1.5, 2.5, 3.5),
        fuse_delays=(3.6, 3.6, 3.6),
    )


def q2_anchor() -> ThreeDeployment:
    return ThreeDeployment(
        heading=0.103436697315,
        speed=138.7915231228,
        release_times=(0.2944979344, 1.2944979344, 2.2944979344),
        fuse_delays=(0.5391236072, 0.5391236072, 0.5391236072),
    )


def build_initial_population(seed: int, population_size: int) -> np.ndarray:
    if population_size < 8:
        raise ValueError("DE 初始种群至少需要八个个体。")
    unit = qmc.LatinHypercube(d=DIMENSION, seed=seed).random(population_size)
    population = np.column_stack((2.0 * np.pi * unit[:, 0], 70.0 + 70.0 * unit[:, 1], unit[:, 2:]))
    anchors = [
        encode_deployment(q2_anchor()),
        encode_deployment(q1_anchor()),
    ]
    rng = np.random.default_rng(seed + 1000)
    for base in (anchors[0], anchors[1]):
        varied = base.copy()
        varied[0] = (varied[0] + rng.normal(0.0, 0.10)) % (2.0 * np.pi)
        varied[1] = np.clip(varied[1] + rng.normal(0.0, 6.0), 70.0, 140.0)
        varied[2:] = np.clip(varied[2:] + rng.normal(0.0, 0.08, 6), 0.0, 1.0)
        anchors.append(varied)
    population[: len(anchors)] = np.asarray(anchors)
    return population


def build_surface_initial_population(
    seed: int,
    population_size: int,
    point_population: np.ndarray,
    point_energies: np.ndarray,
) -> np.ndarray:
    """用点目标全域搜索的多个优质个体初始化完整表面搜索。"""

    population = build_initial_population(seed + 50_000, population_size)
    order = np.argsort(point_energies)
    elite_count = min(8, population_size, len(order))
    elites = np.asarray(point_population[order[:elite_count]], dtype=float)
    population[:elite_count] = elites
    rng = np.random.default_rng(seed + 60_000)
    cursor = elite_count
    while cursor < min(population_size, 3 * elite_count):
        base = elites[(cursor - elite_count) % elite_count].copy()
        base[0] = (base[0] + rng.normal(0.0, 0.08)) % (2.0 * np.pi)
        base[1] = np.clip(base[1] + rng.normal(0.0, 4.0), 70.0, 140.0)
        base[2:] = np.clip(base[2:] + rng.normal(0.0, 0.06, 6), 0.0, 1.0)
        population[cursor] = base
        cursor += 1
    return population


def run_point_search(profile: SearchProfile, seed: int) -> PointSearchResult:
    """用独立随机种子进行一次无锚点的八维点目标全域 DE。"""

    objective = PointTargetObjective(step=0.02)
    result = differential_evolution(
        objective,
        bounds=BOUNDS,
        strategy="best1bin",
        popsize=max(5, profile.point_population_size // DIMENSION),
        maxiter=profile.point_maxiter,
        mutation=(0.5, 1.0),
        recombination=0.9,
        seed=seed,
        init="latinhypercube",
        tol=1.0e-7,
        atol=0.0,
        polish=False,
        updating="immediate",
        workers=1,
    )
    population = np.asarray(result.population, dtype=float)
    energy = np.asarray(result.population_energies, dtype=float)
    best_index = int(np.argmin(energy))
    return PointSearchResult(
        vector=population[best_index].copy(),
        population=population,
        energies=energy,
        evaluations=int(result.nfev),
        message=str(result.message),
    )


def make_candidate(
    vector: Iterable[float], result: JointEvaluationResult, source: str, seed: int | str, rank: int
) -> Candidate:
    values = np.asarray(tuple(float(value) for value in vector), dtype=float)
    return Candidate(values, decode_vector(values), result, source, seed, rank)


def run_seed(profile: SearchProfile, seed: int) -> SearchRun:
    point_started = time.perf_counter()
    point_result = run_point_search(profile, seed)
    point_elapsed = time.perf_counter() - point_started
    angle_count, height_count, radial_count, scan_step = profile.search_grid
    evaluator = UniformJointEvaluator(
        angle_count,
        height_count,
        radial_count,
        scan_step,
        root_tolerance=max(1.0e-5, scan_step * 1.0e-3),
    )
    objective = DurationObjective(evaluator)
    history: list[float] = []

    def callback(intermediate_result) -> bool:
        current = max(0.0, -float(intermediate_result.fun))
        history.append(max(history[-1], current) if history else current)
        return False

    surface_initial = build_surface_initial_population(
        seed,
        profile.population_size,
        point_result.population,
        point_result.energies,
    )
    started = time.perf_counter()
    result = differential_evolution(
        objective,
        bounds=BOUNDS,
        strategy="best1bin",
        maxiter=profile.maxiter,
        popsize=10,
        mutation=(0.5, 1.0),
        recombination=0.9,
        seed=seed,
        init=surface_initial,
        tol=1.0e-5,
        atol=0.0,
        polish=False,
        updating="immediate",
        workers=1,
        callback=callback,
    )
    elapsed = point_elapsed + time.perf_counter() - started
    population = np.asarray(result.population, dtype=float)
    energies = np.asarray(result.population_energies, dtype=float)
    order = np.argsort(energies)
    candidates = tuple(
        make_candidate(
            population[index],
            objective.result_for(population[index]),
            "de_population",
            seed,
            rank,
        )
        for rank, index in enumerate(order[: profile.top_per_seed], start=1)
    )
    best_result = objective.result_for(result.x)
    best = make_candidate(result.x, best_result, "de_best", seed, 0)
    if not history:
        history.append(best_result.duration)
    return SearchRun(
        seed=seed,
        elapsed_seconds=elapsed,
        iterations=int(result.nit),
        evaluations=int(point_result.evaluations + result.nfev),
        success=bool(result.success),
        message=(
            f"point: {point_result.message}; surface: {result.message}"
        ),
        best=best,
        candidates=candidates,
        history=tuple(history),
    )


def candidate_key(candidate: Candidate) -> tuple[float, ...]:
    return tuple(round(float(value), 6) for value in candidate.vector)


def deduplicate_candidates(candidates: Iterable[Candidate]) -> list[Candidate]:
    unique: dict[tuple[float, ...], Candidate] = {}
    for candidate in candidates:
        key = candidate_key(candidate)
        previous = unique.get(key)
        if previous is None or candidate.result.duration > previous.result.duration:
            unique[key] = candidate
    return sorted(unique.values(), key=lambda item: (-item.result.duration, candidate_key(item)))


def rerank_candidates(candidates: Sequence[Candidate], profile: SearchProfile) -> list[Candidate]:
    angle_count, height_count, radial_count, scan_step = profile.rerank_grid
    evaluator = UniformJointEvaluator(
        angle_count,
        height_count,
        radial_count,
        scan_step,
        root_tolerance=max(1.0e-7, scan_step * 1.0e-4),
    )
    reranked = [
        make_candidate(
            candidate.vector,
            evaluator.evaluate(candidate.deployment),
            "rerank",
            candidate.seed,
            rank,
        )
        for rank, candidate in enumerate(candidates[: profile.rerank_count], start=1)
    ]
    return sorted(reranked, key=lambda item: (-item.result.duration, candidate_key(item)))


def independent_statistics(
    deployment: ThreeDeployment, config: AdaptiveSurfaceConfig
) -> tuple[float, tuple[float, ...]]:
    evaluator = build_evaluator(config)
    intervals: list[tuple[float, float]] = []
    durations: list[float] = []
    for release, delay in zip(deployment.release_times, deployment.fuse_delays):
        single = evaluator.evaluate(
            ThreeDeployment(deployment.heading, deployment.speed, (release,), (delay,))
        )
        intervals.extend(single.intervals)
        durations.append(single.duration)
    merged = ThreeSmokeSimulation.merge_intervals(intervals)
    return ThreeSmokeSimulation.total_duration(merged), tuple(durations)


def independent_uniform_statistics(
    deployment: ThreeDeployment,
    evaluator: UniformJointEvaluator,
) -> tuple[float, tuple[float, ...]]:
    intervals: list[tuple[float, float]] = []
    durations: list[float] = []
    for release, delay in zip(deployment.release_times, deployment.fuse_delays):
        single = evaluator.evaluate(
            ThreeDeployment(deployment.heading, deployment.speed, (release,), (delay,))
        )
        intervals.extend(single.intervals)
        durations.append(single.duration)
    merged = ThreeSmokeSimulation.merge_intervals(intervals)
    return ThreeSmokeSimulation.total_duration(merged), tuple(durations)


def intervals_close(
    first: Sequence[tuple[float, float]], second: Sequence[tuple[float, float]], tolerance: float
) -> bool:
    return len(first) == len(second) and all(
        abs(a0 - b0) <= tolerance and abs(a1 - b1) <= tolerance
        for (a0, a1), (b0, b1) in zip(first, second)
    )


def final_review(candidates: Sequence[Candidate], profile: SearchProfile) -> list[FinalReview]:
    angle_count, height_count, radial_count, scan_step = profile.final_grid
    evaluator = UniformJointEvaluator(
        angle_count,
        height_count,
        radial_count,
        scan_step,
        root_tolerance=1.0e-7,
    )
    reviewed: list[FinalReview] = []
    for rank, candidate in enumerate(candidates[: profile.final_count], start=1):
        result = evaluator.evaluate(candidate.deployment)
        independent, individual = independent_uniform_statistics(
            candidate.deployment, evaluator
        )
        reviewed.append(
            FinalReview(
                candidate=make_candidate(candidate.vector, result, "final_uniform", candidate.seed, rank),
                result=result,
                independent_duration=independent,
                individual_durations=individual,
                uniform_review=UniformReview(
                    intervals=candidate.result.intervals,
                    duration=candidate.result.duration,
                    angle_count=profile.rerank_grid[0],
                    height_count=profile.rerank_grid[1],
                    radial_count=profile.rerank_grid[2],
                ),
                dense_uniform_review=evaluator.review_from_result(result),
            )
        )
    reviewed.sort(key=lambda item: (-item.result.duration, candidate_key(item.candidate)))
    if not reviewed:
        return reviewed
    winner = reviewed[0]
    verify_angle, verify_height, verify_radial, verify_step = (
        profile.verification_grid
    )
    verification_evaluator = UniformJointEvaluator(
        verify_angle,
        verify_height,
        verify_radial,
        verify_step,
        root_tolerance=1.0e-8,
    )
    verified_result = verification_evaluator.evaluate(winner.candidate.deployment)
    verified_independent, verified_individual = independent_uniform_statistics(
        winner.candidate.deployment, verification_evaluator
    )
    adaptive_evaluator = build_evaluator(profile.final_adaptive)
    adaptive = adaptive_evaluator.evaluate(
        winner.candidate.deployment, collect_diagnostics=True
    )
    reviewed[0] = FinalReview(
        candidate=make_candidate(
            winner.candidate.vector,
            verified_result,
            "verification_uniform",
            winner.candidate.seed,
            winner.candidate.rank,
        ),
        result=verified_result,
        independent_duration=verified_independent,
        individual_durations=verified_individual,
        uniform_review=winner.uniform_review,
        dense_uniform_review=verification_evaluator.review_from_result(
            verified_result
        ),
        adaptive_review=adaptive,
    )
    return reviewed


def format_intervals(intervals: Sequence[tuple[float, float]]) -> str:
    return "; ".join(f"[{start:.10f}, {end:.10f}]" for start, end in intervals)


def point_columns(prefix: str, point: np.ndarray) -> dict[str, str]:
    return {f"{prefix}_{axis}_m": f"{value:.10f}" for axis, value in zip(("x", "y", "z"), point)}


def candidate_row(candidate: Candidate) -> dict[str, str]:
    deployment, result = candidate.deployment, candidate.result
    row = {
        "source": candidate.source,
        "seed": str(candidate.seed),
        "rank": str(candidate.rank),
        "theta_rad": f"{deployment.heading:.12f}",
        "theta_deg": f"{np.degrees(deployment.heading):.10f}",
        "uav_speed_mps": f"{deployment.speed:.10f}",
        "feasible": str(result.feasible),
        "reason": result.reason,
        "joint_intervals_s": format_intervals(result.intervals),
        "joint_duration_s": f"{result.duration:.10f}",
        "diagnostic_checked_patches": str(result.checked_patches),
        "diagnostic_uncertain_checks": str(result.uncertain_checks),
    }
    for index, (release, delay, burst) in enumerate(
        zip(deployment.release_times, deployment.fuse_delays, deployment.burst_times), start=1
    ):
        point_index = index - 1
        row.update(
            {
                f"release_time_{index}_s": f"{release:.10f}",
                f"fuse_delay_{index}_s": f"{delay:.10f}",
                f"burst_time_{index}_s": f"{burst:.10f}",
            }
        )
        row.update(point_columns(f"release_point_{index}", result.release_points[point_index]))
        row.update(point_columns(f"burst_point_{index}", result.burst_points[point_index]))
    return row


def write_csv(path: Path, rows: Sequence[dict[str, str]]) -> None:
    if not rows:
        raise RuntimeError(f"没有可写入 {path.name} 的记录。")
    fields = list(dict.fromkeys(key for row in rows for key in row))
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def save_outputs(
    profile: SearchProfile,
    runs: Sequence[SearchRun],
    reviewed: Sequence[FinalReview],
    output_dir: Path = Q3_DIR,
) -> tuple[Path, ...]:
    """保存本次搜索已有的结果；分布式子任务可只传入 ``runs``。"""

    output_dir.mkdir(parents=True, exist_ok=True)
    search_rows: list[dict[str, str]] = []
    history_rows: list[dict[str, str]] = []
    for run in runs:
        row = candidate_row(run.best)
        row.update(
            {
                "profile": profile.name,
                "record_type": "de_best",
                "elapsed_seconds": f"{run.elapsed_seconds:.6f}",
                "iterations": str(run.iterations),
                "function_evaluations": str(run.evaluations),
                "success": str(run.success),
                "message": run.message,
            }
        )
        search_rows.append(row)
        for candidate in run.candidates:
            candidate_row_data = candidate_row(candidate)
            candidate_row_data.update({"profile": profile.name, "record_type": "de_population"})
            search_rows.append(candidate_row_data)
        history_rows.extend(
            {
                "profile": profile.name,
                "seed": str(run.seed),
                "generation": str(generation),
                "best_joint_duration_s": f"{duration:.10f}",
            }
            for generation, duration in enumerate(run.history, start=1)
        )
    best_rows: list[dict[str, str]] = []
    diagnostics: list[dict[str, str]] = []
    for rank, item in enumerate(reviewed, start=1):
        row = candidate_row(item.candidate)
        row.update(
            {
                "profile": profile.name,
                "final_rank": str(rank),
                "selected": str(rank == 1),
                "independent_duration_s": f"{item.independent_duration:.10f}",
                "synergy_gain_s": f"{item.result.duration - item.independent_duration:.10f}",
                "individual_durations_s": "; ".join(f"{value:.10f}" for value in item.individual_durations),
                "uniform_duration_s": "" if item.uniform_review is None else f"{item.uniform_review.duration:.10f}",
                "uniform_intervals_s": "" if item.uniform_review is None else format_intervals(item.uniform_review.intervals),
                "dense_uniform_duration_s": "" if item.dense_uniform_review is None else f"{item.dense_uniform_review.duration:.10f}",
                "adaptive_lower_bound_s": "" if item.adaptive_review is None else f"{item.adaptive_review.duration:.10f}",
            }
        )
        best_rows.append(row)
        diagnostics_source = (
            item.adaptive_review.diagnostics
            if item.adaptive_review is not None
            else item.result.diagnostics
        )
        for diagnostic in diagnostics_source:
            diagnostic_row = {
                "final_rank": str(rank),
                "time_s": f"{diagnostic.time:.10f}",
                "status": diagnostic.status,
                "active_indices": ",".join(str(index + 1) for index in diagnostic.active_indices),
                "checked_patches": str(diagnostic.checked_patches),
                "passed_patches": str(diagnostic.passed_patches),
                "uncertain_patches": str(diagnostic.uncertain_patches),
                "witness_distance_m": "" if diagnostic.witness_distance is None else f"{diagnostic.witness_distance:.10f}",
            }
            if diagnostic.witness_point is not None:
                diagnostic_row.update(point_columns("witness_point", diagnostic.witness_point))
            diagnostics.append(diagnostic_row)
    paths: list[Path] = []
    if search_rows:
        search_path = output_dir / "q3_search_runs.csv"
        write_csv(search_path, search_rows)
        paths.append(search_path)
    if history_rows:
        history_path = output_dir / "q3_de_history.csv"
        write_csv(history_path, history_rows)
        paths.append(history_path)
    if best_rows:
        best_path = output_dir / "q3_best_solution.csv"
        diagnostics_path = output_dir / "q3_surface_diagnostics.csv"
        write_csv(best_path, best_rows)
        write_csv(diagnostics_path, diagnostics or [{"status": "no_final_candidate"}])
        paths.extend((best_path, diagnostics_path))
    if reviewed:
        paths.append(write_result_workbook(reviewed[0], output_dir))
    return tuple(paths)


def normalize_header(value: object) -> str:
    return "".join(str(value or "").split()).replace("：", ":")


def header_column(worksheet, keywords: Sequence[str]) -> int:
    for cell in worksheet[1]:
        value = normalize_header(cell.value)
        if all(keyword in value for keyword in keywords):
            return cell.column
    raise KeyError(f"模板缺少列：{'/'.join(keywords)}")


def write_result_workbook(review: FinalReview, output_dir: Path = Q3_DIR) -> Path:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"未找到结果模板：{TEMPLATE_PATH}")
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / "result1.xlsx"
    shutil.copy2(TEMPLATE_PATH, output)
    workbook = load_workbook(output)
    worksheet = workbook.active
    columns = {
        "heading": header_column(worksheet, ("无人机", "方向")),
        "speed": header_column(worksheet, ("无人机", "速度")),
        "number": header_column(worksheet, ("烟幕", "编号")),
        "release_x": header_column(worksheet, ("投放点", "x")),
        "release_y": header_column(worksheet, ("投放点", "y")),
        "release_z": header_column(worksheet, ("投放点", "z")),
        "burst_x": header_column(worksheet, ("起爆点", "x")),
        "burst_y": header_column(worksheet, ("起爆点", "y")),
        "burst_z": header_column(worksheet, ("起爆点", "z")),
        "duration": header_column(worksheet, ("有效", "时长")),
    }
    deployment, result = review.candidate.deployment, review.result
    for index in range(3):
        row = index + 2
        worksheet.cell(row, columns["heading"]).value = float(np.degrees(deployment.heading))
        worksheet.cell(row, columns["speed"]).value = float(deployment.speed)
        worksheet.cell(row, columns["number"]).value = index + 1
        release = result.release_points[index]
        burst = result.burst_points[index]
        worksheet.cell(row, columns["release_x"]).value = float(release[0])
        worksheet.cell(row, columns["release_y"]).value = float(release[1])
        worksheet.cell(row, columns["release_z"]).value = float(release[2])
        worksheet.cell(row, columns["burst_x"]).value = float(burst[0])
        worksheet.cell(row, columns["burst_y"]).value = float(burst[1])
        worksheet.cell(row, columns["burst_z"]).value = float(burst[2])
        worksheet.cell(row, columns["duration"]).value = float(review.individual_durations[index])
    workbook.save(output)
    return output


def solve(profile: SearchProfile) -> tuple[list[SearchRun], list[FinalReview]]:
    runs: list[SearchRun] = []
    for seed in profile.seeds:
        print(f"正在运行 DE：档位={profile.name}，随机种子={seed}...", flush=True)
        run = run_seed(profile, seed)
        runs.append(run)
        print(
            f"  完成：{run.elapsed_seconds:.1f} s，"
            f"低精度联合遮蔽={run.best.result.duration:.6f} s。",
            flush=True,
        )
    pool = deduplicate_candidates(
        candidate for run in runs for candidate in (run.best, *run.candidates)
    )
    print("正在执行中精度重排与最终完整表面复核...", flush=True)
    reranked = rerank_candidates(pool, profile)
    return runs, final_review(reranked, profile)


def print_final_result(reviewed: Sequence[FinalReview]) -> None:
    if not reviewed:
        raise RuntimeError("没有候选方案通过最终联合评价。")
    result = reviewed[0]
    deployment = result.candidate.deployment
    print("\n问题三推荐方案（分层完整圆柱表面联合求解）")
    print(f"  航向角: {deployment.heading:.10f} rad / {np.degrees(deployment.heading):.8f}°")
    print(f"  飞行速度: {deployment.speed:.8f} m/s")
    for index, (release, delay, burst) in enumerate(
        zip(deployment.release_times, deployment.fuse_delays, deployment.burst_times), start=1
    ):
        print(f"  弹 {index}: 投放 {release:.8f} s，引爆延迟 {delay:.8f} s，起爆 {burst:.8f} s")
        print(f"       投放点 {np.array2string(result.result.release_points[index - 1], precision=5)}")
        print(f"       起爆点 {np.array2string(result.result.burst_points[index - 1], precision=5)}")
    print(f"  联合遮蔽区间: {format_intervals(result.result.intervals)} s")
    print(f"  联合遮蔽总时长: {result.result.duration:.10f} s")
    print(f"  独立遮蔽并集时长: {result.independent_duration:.10f} s")
    print(f"  协同增益: {result.result.duration - result.independent_duration:.10f} s")
    if result.uniform_review is not None:
        grid = result.uniform_review
        print(
            f"  {grid.angle_count}×{grid.height_count}×{grid.radial_count} "
            f"候选重排: {grid.duration:.10f} s"
        )
    if result.dense_uniform_review is not None:
        grid = result.dense_uniform_review
        print(
            f"  {grid.angle_count}×{grid.height_count}×{grid.radial_count} "
            f"最终完整表面复核: {grid.duration:.10f} s"
        )
    if result.adaptive_review is not None:
        print(
            "  严格自适应保守下界: "
            f"{result.adaptive_review.duration:.10f} s"
        )


def parse_seed_list(text: str) -> tuple[int, ...]:
    """解析 ``20250831,20250832`` 形式的独立 DE 种子列表。"""

    try:
        seeds = tuple(int(item.strip()) for item in text.split(",") if item.strip())
    except ValueError as error:
        raise argparse.ArgumentTypeError("随机种子必须是逗号分隔的整数。") from error
    if not seeds or len(set(seeds)) != len(seeds):
        raise argparse.ArgumentTypeError("至少提供一个且不能重复的随机种子。")
    return seeds


def main() -> None:
    parser = argparse.ArgumentParser(description="问题三三烟幕联合遮蔽标准 DE 求解")
    parser.add_argument("--profile", choices=tuple(PROFILES), default="standard")
    parser.add_argument(
        "--seeds",
        type=parse_seed_list,
        help="仅运行指定的独立 DE 种子，例如 20250831,20250832；用于多人并行。",
    )
    parser.add_argument(
        "--search-only",
        action="store_true",
        help="只运行 DE 并保存候选/历史，不进行重复的最终复核。",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="结果目录；多人并行时必须为每人指定不同目录。",
    )
    arguments = parser.parse_args()
    profile = PROFILES[arguments.profile]
    if arguments.seeds is not None:
        invalid = set(arguments.seeds) - set(profile.seeds)
        if invalid:
            parser.error(f"指定种子不属于 {profile.name} 档：{sorted(invalid)}")
        profile = replace(profile, seeds=arguments.seeds)
    output_dir = arguments.output_dir or Q3_DIR
    started = time.perf_counter()
    if arguments.search_only:
        runs = [run_seed(profile, seed) for seed in profile.seeds]
        reviewed: list[FinalReview] = []
    else:
        runs, reviewed = solve(profile)
        print_final_result(reviewed)
    for path in save_outputs(profile, runs, reviewed, output_dir):
        print(f"结果已保存至: {path}")
    print(f"总运行时间: {time.perf_counter() - started:.2f} s")


if __name__ == "__main__":
    main()

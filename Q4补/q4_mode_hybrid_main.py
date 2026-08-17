"""Independent Q4 experiment: classified seeds + LHS + one joint DE objective."""

from __future__ import annotations

import argparse
import csv
from dataclasses import asdict
import math
from pathlib import Path
import shutil
import sys
import time
from typing import Iterable, Sequence

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import differential_evolution

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from Q4 import q4_main as base  # noqa: E402
from mode_seed_factory import MODE_LABELS, PopulationBundle, PopulationRecord, make_baseline_population, make_mode_hybrid_population  # noqa: E402


class BatchedJointEvaluator(base.JointEvaluator):
    """Memory-safe wrapper around the original evaluator.

    The original vectorised evaluator is fast for coarse grids, but a complete
    high-density evaluation can contain thousands of time points.  Splitting
    only the time axis preserves every formula and every surface sample while
    avoiding a multi-gigabyte ``time × surface × coordinate`` temporary array.
    """

    def __init__(self, *args, time_batch_size: int = 8, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        if time_batch_size < 1:
            raise ValueError("time_batch_size must be positive")
        self.time_batch_size = int(time_batch_size)

    def status_batch(
        self, strategy: base.Strategy, times: np.ndarray, active: tuple[int, ...]
    ) -> np.ndarray:
        values = np.asarray(times, dtype=float)
        if len(values) <= self.time_batch_size:
            return super().status_batch(strategy, values, active)
        return np.concatenate(
            [
                super().status_batch(strategy, values[start : start + self.time_batch_size], active)
                for start in range(0, len(values), self.time_batch_size)
            ]
        )


def parse_seeds(text: str) -> tuple[int, ...]:
    try:
        values = tuple(int(item.strip()) for item in text.split(",") if item.strip())
    except ValueError as error:
        raise argparse.ArgumentTypeError("Seeds must be comma-separated integers.") from error
    if not values:
        raise argparse.ArgumentTypeError("At least one seed is required.")
    return values


def write_csv(path: Path, rows: Sequence[dict[str, object]]) -> None:
    if not rows:
        raise RuntimeError(f"No rows available for {path.name}.")
    fields = list(dict.fromkeys(key for row in rows for key in row))
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def format_intervals(intervals: Iterable[tuple[float, float]]) -> str:
    return "; ".join(f"[{left:.10f}, {right:.10f}]" for left, right in intervals)


def convergence_checkpoint_rows(
    init_strategy: str,
    profile: base.SearchProfile,
    runs: Sequence[base.SearchRun],
    checkpoints: Sequence[int],
    stability_threshold: float,
) -> list[dict[str, object]]:
    """Extract predeclared generations without altering the optimiser itself."""
    rows: list[dict[str, object]] = []
    for run in runs:
        previous_generation: int | None = None
        previous_duration: float | None = None
        for generation in sorted(set(int(item) for item in checkpoints)):
            if generation < 1 or generation > len(run.history):
                continue
            duration = float(run.history[generation - 1])
            increment = None if previous_duration is None else duration - previous_duration
            rows.append(
                {
                    "init_strategy": init_strategy,
                    "profile": profile.name,
                    "seed": run.seed,
                    "checkpoint_generation": generation,
                    "best_duration_s": duration,
                    "previous_checkpoint": previous_generation,
                    "increment_s": increment,
                    "stability_threshold_s": stability_threshold,
                    "stable_from_previous": None if increment is None else increment <= stability_threshold + 1.0e-12,
                }
            )
            previous_generation, previous_duration = generation, duration
    return rows


def write_convergence_summary(
    path: Path,
    rows: Sequence[dict[str, object]],
    final_generation: int,
    stability_threshold: float,
) -> None:
    final_rows = [row for row in rows if int(row["checkpoint_generation"]) == final_generation]
    if not final_rows:
        statement = "No requested checkpoint was reached; no stability conclusion is made."
    elif all(bool(row["stable_from_previous"]) for row in final_rows):
        statement = "All seeds satisfy the practical stability threshold on the final checkpoint interval."
    else:
        statement = "At least one seed still exceeds the practical stability threshold on the final checkpoint interval."
    lines = [
        "# Q4 convergence checkpoint summary",
        "",
        f"- Stability threshold: `{stability_threshold:.6f}` s improvement per checkpoint interval.",
        f"- Final checkpoint: generation `{final_generation}`.",
        f"- Conclusion: {statement}",
        "",
        "The detailed per-seed values are stored in `q4_convergence_checkpoints.csv`.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _record_row(record: PopulationRecord) -> dict[str, object]:
    row: dict[str, object] = {
        "seed": record.seed,
        "index": record.index,
        "source": record.source,
        "requested_mode": record.requested_mode,
        "actual_initial_class": record.classification,
        "fallback": record.fallback,
        "feasible": base.validate_strategy(record.strategy) is None,
    }
    for index, plan in enumerate(record.strategy.plans, start=1):
        row.update(
            {
                f"theta_{index}_deg": np.degrees(plan.heading),
                f"speed_{index}_mps": plan.speed,
                f"release_{index}_s": plan.release_time,
                f"delay_{index}_s": plan.fuse_delay,
                f"burst_{index}_s": plan.burst_time,
            }
        )
    return row


def _summary_rows(bundle: PopulationBundle) -> list[dict[str, object]]:
    rows = []
    for requested in sorted({record.requested_mode for record in bundle.records}):
        group = [record for record in bundle.records if record.requested_mode == requested]
        spans = [max(plan.burst_time for plan in record.strategy.plans) - min(plan.burst_time for plan in record.strategy.plans) for record in group]
        rows.append(
            {
                "seed": group[0].seed,
                "requested_mode": requested,
                "label": MODE_LABELS.get(requested, requested),
                "requested_count": bundle.requested_counts.get(requested, len(group)),
                "actual_count": len(group),
                "fallback_count": sum(record.fallback for record in group),
                "mean_burst_span_s": float(np.mean(spans)),
                "min_burst_span_s": float(np.min(spans)),
                "max_burst_span_s": float(np.max(spans)),
            }
        )
    return rows


def build_population(
    init_strategy: str,
    profile: base.SearchProfile,
    libraries: Sequence[Sequence[base.SingleCandidate]],
    seed: int,
) -> PopulationBundle:
    if init_strategy == "mode-hybrid":
        return make_mode_hybrid_population(profile, libraries, seed)
    if init_strategy == "baseline":
        return make_baseline_population(profile, libraries, seed)
    raise ValueError(f"Unknown initialisation strategy: {init_strategy}")


def run_joint_seed(
    profile: base.SearchProfile,
    bundle: PopulationBundle,
    seed: int,
    workers: int,
    time_batch_size: int,
) -> base.SearchRun:
    evaluator = BatchedJointEvaluator(*profile.search_grid, time_batch_size=time_batch_size)
    objective = base.Objective(evaluator, (0, 1, 2))
    history: list[float] = []

    def callback(vector: np.ndarray, convergence: float) -> bool:
        del convergence
        duration = objective.result_for(vector).duration
        history.append(max(history[-1], duration) if history else duration)
        return False

    started = time.perf_counter()
    result = differential_evolution(
        objective,
        bounds=base.JOINT_BOUNDS,
        strategy="best1bin",
        init=bundle.population,
        maxiter=profile.maxiter,
        mutation=(0.5, 1.0),
        recombination=0.9,
        seed=seed,
        polish=False,
        tol=1.0e-5,
        atol=0.0,
        callback=callback,
        updating="immediate" if workers == 1 else "deferred",
        workers=workers,
    )
    elapsed = time.perf_counter() - started
    population = np.asarray(result.population, dtype=float)
    energies = np.asarray(result.population_energies, dtype=float)
    order = np.argsort(energies)[: profile.top_per_seed]
    candidates = tuple(
        base.JointCandidate(population[index].copy(), objective.result_for(population[index]), "mode_hybrid_population", seed)
        for index in order
    )
    best = base.JointCandidate(np.asarray(result.x, dtype=float), objective.result_for(result.x), "mode_hybrid_best", seed)
    return base.SearchRun(
        seed=seed,
        elapsed_seconds=elapsed,
        iterations=int(result.nit),
        evaluations=int(result.nfev),
        success=bool(result.success),
        message=str(result.message),
        best=best,
        candidates=candidates,
        history=tuple(history or [best.result.duration]),
    )


def rerank(
    candidates: Sequence[base.JointCandidate],
    grid: tuple[int, int, int, float],
    count: int,
    source: str,
    time_batch_size: int,
) -> list[base.JointCandidate]:
    evaluator = BatchedJointEvaluator(*grid, time_batch_size=time_batch_size)
    reviewed = [
        base.JointCandidate(
            candidate.encoded,
            evaluator.evaluate(base.decode_strategy(candidate.encoded)),
            source,
            candidate.seed,
        )
        for candidate in candidates[:count]
    ]
    return sorted(reviewed, key=lambda item: -item.result.duration)


def write_workbook(output_dir: Path, evaluation: base.Evaluation, individual: Sequence[float]) -> Path:
    template = PROJECT_ROOT / "Resources" / "result2.xlsx"
    output = output_dir / "result2.xlsx"
    shutil.copy2(template, output)
    workbook = load_workbook(output)
    worksheet = workbook.active
    for index, plan in enumerate(evaluation.strategy.plans, start=2):
        release = evaluation.release_points[index - 2]
        burst = evaluation.burst_points[index - 2]
        values = (
            plan.name,
            float(np.degrees(plan.heading)),
            plan.speed,
            float(release[0]),
            float(release[1]),
            float(release[2]),
            float(burst[0]),
            float(burst[1]),
            float(burst[2]),
            float(individual[index - 2]),
        )
        for column, value in enumerate(values, start=1):
            worksheet.cell(index, column).value = value
    workbook.save(output)
    return output


def write_summary(
    path: Path,
    init_strategy: str,
    profile: base.SearchProfile,
    verification: base.Evaluation,
    independent: float,
    runs: Sequence[base.SearchRun],
) -> None:
    gain = verification.duration - independent
    mode = "时间接力为主" if abs(gain) <= 1.0e-6 else "存在可观测的空间协同"
    lines = [
        "# Q4 分类种子混合搜索运行摘要",
        "",
        f"- 初始化策略：`{init_strategy}`",
        f"- 搜索网格：`{profile.search_grid[0]}×{profile.search_grid[1]}×{profile.search_grid[2]}`，时间步长 `{profile.search_grid[3]}` s",
        f"- 最终复核网格：`{verification.grid[0]}×{verification.grid[1]}×{verification.grid[2]}`，时间步长 `{verification.grid[3]}` s",
        f"- 联合完整遮蔽总时长：`{verification.duration:.10f}` s",
        f"- 单弹完整遮蔽区间并集：`{independent:.10f}` s",
        f"- 空间协同增益：`{gain:.10f}` s",
        f"- 事后结构解释：{mode}。",
        "",
        "## 多种子搜索",
        "",
        "| Seed | 搜索阶段时长 (s) | 迭代代数 | 耗时 (s) |",
        "|---:|---:|---:|---:|",
    ]
    lines.extend(f"| {run.seed} | {run.best.result.duration:.6f} | {run.iterations} | {run.elapsed_seconds:.1f} |" for run in runs)
    lines.extend(
        [
            "",
            "分类标签仅用于构造初始种群；DE 与最终复核始终使用同一联合完整遮蔽目标，因此不对最终方案施加接力、重叠或空间分工等硬约束。",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def save_outputs(
    output_dir: Path,
    init_strategy: str,
    profile: base.SearchProfile,
    bundles: Sequence[PopulationBundle],
    runs: Sequence[base.SearchRun],
    verification: base.Evaluation,
    independent: float,
    individual: tuple[float, ...],
    independent_intervals: tuple[tuple[float, float], ...],
    checkpoints: tuple[int, ...] = (600, 750, 900),
    stability_threshold: float = 0.02,
) -> None:
    initial_rows = [_record_row(record) for bundle in bundles for record in bundle.records]
    summary_rows = [row for bundle in bundles for row in _summary_rows(bundle)]
    search_rows = [
        {
            "init_strategy": init_strategy,
            "profile": profile.name,
            "seed": run.seed,
            "elapsed_seconds": run.elapsed_seconds,
            "iterations": run.iterations,
            "function_evaluations": run.evaluations,
            "success": run.success,
            "message": run.message,
            "search_duration_s": run.best.result.duration,
            "intervals_s": format_intervals(run.best.result.intervals),
        }
        for run in runs
    ]
    history_rows = [
        {
            "init_strategy": init_strategy,
            "profile": profile.name,
            "seed": run.seed,
            "generation": generation,
            "best_duration_s": duration,
        }
        for run in runs
        for generation, duration in enumerate(run.history, start=1)
    ]
    checkpoint_rows = convergence_checkpoint_rows(
        init_strategy, profile, runs, checkpoints, stability_threshold
    )
    best_rows = []
    for index, plan in enumerate(verification.strategy.plans):
        release, burst = verification.release_points[index], verification.burst_points[index]
        best_rows.append(
            {
                "init_strategy": init_strategy,
                "profile": profile.name,
                "uav": plan.name,
                "theta_deg": np.degrees(plan.heading),
                "speed_mps": plan.speed,
                "release_time_s": plan.release_time,
                "fuse_delay_s": plan.fuse_delay,
                "burst_time_s": plan.burst_time,
                "release_x_m": release[0],
                "release_y_m": release[1],
                "release_z_m": release[2],
                "burst_x_m": burst[0],
                "burst_y_m": burst[1],
                "burst_z_m": burst[2],
                "individual_duration_s": individual[index],
                "joint_duration_s": verification.duration,
                "joint_intervals_s": format_intervals(verification.intervals),
                "independent_union_duration_s": independent,
                "independent_union_intervals_s": format_intervals(independent_intervals),
                "synergy_gain_s": verification.duration - independent,
                "verification_grid": f"{verification.grid[0]}x{verification.grid[1]}x{verification.grid[2]}@{verification.grid[3]}",
            }
        )
    write_csv(output_dir / "q4_initial_population.csv", initial_rows)
    write_csv(output_dir / "q4_mode_seed_summary.csv", summary_rows)
    write_csv(output_dir / "q4_search_runs.csv", search_rows)
    write_csv(output_dir / "q4_de_history.csv", history_rows)
    if checkpoint_rows:
        write_csv(output_dir / "q4_convergence_checkpoints.csv", checkpoint_rows)
        write_convergence_summary(
            output_dir / "q4_convergence_summary.md",
            checkpoint_rows,
            max(checkpoints),
            stability_threshold,
        )
    write_csv(output_dir / "q4_best_solution.csv", best_rows)
    write_workbook(output_dir, verification, individual)
    write_summary(output_dir / "run_summary.md", init_strategy, profile, verification, independent, runs)


def ensure_output_dir(path: Path, overwrite: bool) -> None:
    expected = ("q4_initial_population.csv", "q4_best_solution.csv", "result2.xlsx", "run_summary.md")
    if path.exists() and not overwrite and any((path / name).exists() for name in expected):
        raise FileExistsError(f"{path} already contains a run. Use --overwrite or a new --output-dir.")
    path.mkdir(parents=True, exist_ok=True)


def run(
    profile: base.SearchProfile,
    init_strategy: str,
    output_dir: Path,
    workers: int,
    time_batch_size: int,
    check_only: bool,
    checkpoints: tuple[int, ...],
    stability_threshold: float,
) -> None:
    base.run_regression_checks()
    libraries = tuple(base.build_single_library(profile, index, 20250900 + index) for index in range(3))
    bundles = [build_population(init_strategy, profile, libraries, seed) for seed in profile.seeds]
    if any(bundle.population.shape != (profile.population_size, base.DIMENSION) for bundle in bundles):
        raise AssertionError("Initial population shape is incorrect.")
    if check_only:
        print("Check passed: Q1 regression and all initial populations are feasible.")
        return
    runs = [
        run_joint_seed(profile, bundle, seed, workers, time_batch_size)
        for bundle, seed in zip(bundles, profile.seeds)
    ]
    pool = base.deduplicate(candidate for run_item in runs for candidate in (run_item.best, *run_item.candidates))
    middle = rerank(pool, profile.rerank_grid, profile.rerank_count, "middle_rerank", time_batch_size)
    final_candidates = rerank(middle, profile.final_grid, profile.final_count, "final_review", time_batch_size)
    if not final_candidates:
        raise RuntimeError("No candidate survived final complete-surface evaluation.")
    verifier = BatchedJointEvaluator(*profile.verification_grid, time_batch_size=time_batch_size)
    verification = verifier.evaluate(final_candidates[0].result.strategy)
    independent, individual, independent_intervals = base.independent_statistics(verification.strategy, verifier)
    save_outputs(
        output_dir, init_strategy, profile, bundles, runs, verification, independent, individual,
        independent_intervals, checkpoints, stability_threshold,
    )
    print(f"Completed {init_strategy}: T_joint={verification.duration:.10f} s")
    print(f"Outputs: {output_dir}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Independent Q4 mode-hybrid DE experiment")
    parser.add_argument("--profile", choices=tuple(base.PROFILES), default="standard")
    parser.add_argument("--seeds", type=parse_seeds)
    parser.add_argument(
        "--maxiter",
        type=int,
        help="Override the profile's DE generation limit without changing any other search setting.",
    )
    parser.add_argument("--init-strategy", choices=("baseline", "mode-hybrid"), default="mode-hybrid")
    parser.add_argument("--output-dir", default="Q4补/runs/mode_hybrid_standard")
    parser.add_argument("--workers", type=int, default=1)
    parser.add_argument(
        "--time-batch-size",
        type=int,
        default=8,
        help="Maximum number of time samples evaluated together; lower values use less memory.",
    )
    parser.add_argument(
        "--checkpoints",
        type=parse_seeds,
        default=(600, 750, 900),
        help="Comma-separated generations for convergence reporting; defaults to 600,750,900.",
    )
    parser.add_argument(
        "--stability-threshold",
        type=float,
        default=0.02,
        help="Maximum improvement (s) over one checkpoint interval treated as practically stable.",
    )
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--check-only", action="store_true")
    args = parser.parse_args()
    if args.workers == 0 or args.workers < -1:
        parser.error("workers must be -1 or a positive integer")
    if args.time_batch_size < 1:
        parser.error("time-batch-size must be positive")
    if args.stability_threshold < 0.0:
        parser.error("stability-threshold must be non-negative")
    if args.maxiter is not None and args.maxiter < 1:
        parser.error("maxiter must be positive")
    profile = base.PROFILES[args.profile]
    if args.seeds is not None:
        profile = base.replace(profile, seeds=args.seeds)
    if args.maxiter is not None:
        profile = base.replace(profile, maxiter=args.maxiter)
    output_dir = Path(args.output_dir)
    if not output_dir.is_absolute():
        output_dir = PROJECT_ROOT / output_dir
    if not args.check_only:
        ensure_output_dir(output_dir, args.overwrite)
    checkpoints = tuple(sorted(set(int(value) for value in args.checkpoints)))
    run(
        profile, args.init_strategy, output_dir, args.workers, args.time_batch_size, args.check_only,
        checkpoints, args.stability_threshold,
    )


if __name__ == "__main__":
    main()
